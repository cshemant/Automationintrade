import os
import re
import time
import warnings
import argparse
from io import StringIO

os.environ["MPLBACKEND"] = "Agg"

import matplotlib
matplotlib.use("Agg", force=True)
import matplotlib.pyplot as plt
from PIL import Image, ImageDraw, ImageFont
import numpy as np
import pandas as pd
import requests
import yfinance as yf
from bs4 import BeautifulSoup
from sklearn.linear_model import Ridge
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# =========================================================
# DATE + LOCAL STOCK LIST
# =========================================================
# DATE filled = read stock list from local Excel only.
# Expected file for DATE="05/05/2026":
# ../ResultStockList/ResultCalendarStocks/ResultCalendar_05_05_2026.xlsx
DATE = ""#22/05/2026"
StockList = r"../ResultStockList/ResultCalendarStocks/ResultCalendar_{DD_MM_YYYY}.xlsx"

# If a row has blank / nan Symbol, it will be skipped.
# No Moneycontrol / NSE / dynamic web mapping is used.
SKIP_ROWS_WITH_BLANK_SYMBOL = True
CONTINUE_ON_STOCK_ERROR = True

# =========================================================
# SCREENER FETCH BUTTONS
# =========================================================
SCREENER_REQUEST_SLEEP_SECONDS = 2.5
SCREENER_429_RETRY_SECONDS = 12
SCREENER_MAX_RETRIES = 4

# =========================================================
# ORIGINAL RESULT IMPACT SUMMARY CALCULATION BUTTONS
# =========================================================
# Required for summary rows: Direction, Expected Move 1D, Expected Move 3D, Confidence.
EVENT_WINDOWS = [1, 3]
MIN_TRAIN_EVENTS = 6
PRICE_BUFFER_DAYS = 15
FALLBACK_EVENT_OFFSET_DAYS = 20
SHOW_MOVE = "YES"   # YES = show % values, NO = hide as **

# =========================================================
# OUTPUT TOGGLES / IMAGE BUTTONS
# =========================================================
GENERATE_GENERAL_IMAGE = True
GENERATE_INSTAGRAM_IMAGE = True
GENERATE_EXCEL_REPORT = True

GENERAL_IMAGE_OUTPUT_TEMPLATE = r"Image/General/{SYMBOL}.jpeg"
INSTAGRAM_IMAGE_OUTPUT_TEMPLATE = r"Image/Instagram/{SYMBOL}.jpeg"
EXCEL_OUTPUT_TEMPLATE = r"Result/{SYMBOL}_{QUARTER_SAFE}_Result_Analysis.xlsx"


# =========================================================
# SCORECARD VIDEO FRAMES BUTTONS
# =========================================================
# Generates 4 score-overlay frames from the latest General scorecard image.
# Output folders required by user:
#   Image/VideoFrames/General/
#   Image/VideoFrames/Instagram/
GENERATE_SCORECARD_VIDEO_FRAMES = True
GENERATE_SCORECARD_GENERAL_VIDEO_FRAMES = True
GENERATE_SCORECARD_INSTAGRAM_VIDEO_FRAMES = True
SCORECARD_GENERAL_VIDEO_FRAMES_OUTPUT_TEMPLATE = r"Image/VideoFrames/General/{SYMBOL}_Frame_{FRAME_NO}.jpeg"
SCORECARD_INSTAGRAM_VIDEO_FRAMES_OUTPUT_TEMPLATE = r"Image/VideoFrames/Instagram/{SYMBOL}_Frame_{FRAME_NO}.jpeg"
# Backward-compatible fallback. If an older call does not pass a template, General is used.
SCORECARD_VIDEO_FRAMES_OUTPUT_TEMPLATE = SCORECARD_GENERAL_VIDEO_FRAMES_OUTPUT_TEMPLATE
SCORECARD_VIDEO_FRAMES_COUNT = 4

# The full General image remains unchanged. Only the center score text grows
# frame-by-frame, creating a clean zoom-in effect without cropping the scorecard.
SCORECARD_CENTER_SCORE_FONT_SIZES = [46, 58, 72, 88]
SCORECARD_CENTER_SCORE_PREFIX = "Score: "
SCORECARD_CENTER_SCORE_Y = 0.50
SCORECARD_CENTER_BOX_WIDTH = 0.58
SCORECARD_CENTER_BOX_HEIGHT = 0.16
SCORECARD_CENTER_BOX_ALPHA = 190
SCORECARD_CENTER_BOX_RADIUS = 28
SCORECARD_CENTER_BOX_BORDER_WIDTH = 3
SCORECARD_CENTER_SCORE_COLOR = "#FFFFFF"
SCORECARD_CENTER_LABEL_COLOR = "#E5C26A"
# YES = center score text + border use the same condition-based color as the top result strip.
# NO  = use SCORECARD_CENTER_SCORE_COLOR / SCORECARD_CENTER_LABEL_COLOR / SCORECARD_CENTER_BOX_BORDER.
SCORECARD_CENTER_USE_RESULT_COLOR = True
SCORECARD_CENTER_BOX_FILL = "#0B1E3A"
SCORECARD_CENTER_BOX_BORDER = "#46A36A"
SCORECARD_VIDEO_FRAME_QUALITY = 95


# =========================================================
# SUMMARY IMAGE OUTPUT TOGGLES / BUTTONS
# =========================================================
# These generate {SYMBOL}_Summary.jpeg in four brand-aligned variations.
GENERATE_GENERAL_SUMMARY_IMAGE = True
GENERATE_INSTAGRAM_SUMMARY_IMAGE = True
GENERATE_STANDARD_SUMMARY_IMAGE = True
GENERATE_REELS_SUMMARY_IMAGE = True

GENERAL_SUMMARY_IMAGE_OUTPUT_TEMPLATE = r"Image/General/{SYMBOL}_Summary.jpeg"
INSTAGRAM_SUMMARY_IMAGE_OUTPUT_TEMPLATE = r"Image/Instagram/{SYMBOL}_Summary.jpeg"
STANDARD_SUMMARY_IMAGE_OUTPUT_TEMPLATE = r"Image/Standard/{SYMBOL}_Summary.jpeg"
REELS_SUMMARY_IMAGE_OUTPUT_TEMPLATE = r"Image/Reels/{SYMBOL}_Summary.jpeg"


# =========================================================
# INSTAGRAM REEL FRAME PACK BUTTONS
# =========================================================
# Purpose: create multiple light, fast-changing 9:16 frames from the same
# summary payload so Reels do not feel like one heavy image post.
GENERATE_REELS_FRAME_PACK = True
REELS_FRAME_PACK_OUTPUT_TEMPLATE = r"Image/ReelsFrames/{SYMBOL}/Frame_{FRAME_NO}_{FRAME_NAME}.jpeg"

# Reel frame size is true 9:16. At 160 DPI this exports around 1080x1920 px.
REELS_FRAME_FIG_W = 6.75
REELS_FRAME_FIG_H = 12.0
REELS_FRAME_DPI = 160
REELS_FRAME_SHOW_HANDLE = True
REELS_FRAME_HANDLE_TEXT = "automationintrade"
REELS_FRAME_FOOTER_Y = 0.070
REELS_FRAME_DISCLAIMER_TEXT = "Data-based estimate only. Not investment advice."
REELS_FRAME_DISCLAIMER_Y = 0.105

# First frame hook. Keep this very clear; viewer should understand the Reel
# before the voice-over starts.
REELS_FRAME_01_HOOK_TITLE = "RESULT IMPACT IN 5 SECONDS"
REELS_FRAME_01_HOOK_SUBTITLE = "Before you react, check the likely result mood."

# Optional frame duration hints. These are printed only in the console so you
# can quickly edit the Reel timeline.
REELS_FRAME_DURATION_HINT_SECONDS = {
    "Frame_01_Hook": 1.2,
    "Frame_02_Impact": 1.4,
    "Frame_03_Drivers": 1.6,
    "Frame_04_Setup": 1.4,
}

# Summary - General
GENERAL_SUMMARY_FIG_W = 7.5
GENERAL_SUMMARY_FIG_H = 12.0
GENERAL_SUMMARY_DPI = 160
GENERAL_SUMMARY_CARD_X = 0.06
GENERAL_SUMMARY_CARD_Y = 0.16
GENERAL_SUMMARY_CARD_W = 0.88
GENERAL_SUMMARY_CARD_H = 0.66
GENERAL_SUMMARY_TITLE_Y = 0.902
GENERAL_SUMMARY_SUBTITLE_Y = 0.866
GENERAL_SUMMARY_FOOTER_Y = 0.099
GENERAL_SUMMARY_TITLE_SIZE = 26
GENERAL_SUMMARY_SUBTITLE_SIZE = 16
GENERAL_SUMMARY_STRIP_TEXT_SIZE = 20
GENERAL_SUMMARY_SCORE_SIZE = 42
GENERAL_SUMMARY_LABEL_SIZE = 13
GENERAL_SUMMARY_VALUE_SIZE = 15
GENERAL_SUMMARY_PARAGRAPH_SIZE = 13
GENERAL_SUMMARY_LINE_SPACING = 1.20
GENERAL_SUMMARY_STRIP_ROW_GAP = 0.032  # Gap between top result strip and first summary row
GENERAL_SUMMARY_ROW_GAP = 0.017        # Gap/padding between summary rows
GENERAL_SUMMARY_TOP_STRIP_LABEL = "EXPECTED PATTERN"  # Top green/orange/red strip text; keeps score badge from overlapping
GENERAL_SUMMARY_DISCLAIMER_TEXT = "Disclaimer: Data-based estimate only. Not investment advice."
GENERAL_SUMMARY_DISCLAIMER_SIZE = 9
GENERAL_SUMMARY_DISCLAIMER_Y_SHIFT = -0.013
GENERAL_SUMMARY_SHADOW_ALPHA = 0.12
GENERAL_SUMMARY_SHOW_HANDLE = True
GENERAL_SUMMARY_HANDLE_TEXT = "automationintrade"

# Summary - Instagram
INSTAGRAM_SUMMARY_FIG_W = 8.0
INSTAGRAM_SUMMARY_FIG_H = 10.0
INSTAGRAM_SUMMARY_DPI = 160
INSTAGRAM_SUMMARY_CARD_X = 0.085
INSTAGRAM_SUMMARY_CARD_Y = 0.16
INSTAGRAM_SUMMARY_CARD_W = 0.83
INSTAGRAM_SUMMARY_CARD_H = 0.626
INSTAGRAM_SUMMARY_TITLE_Y = 0.870
INSTAGRAM_SUMMARY_SUBTITLE_Y = 0.820
INSTAGRAM_SUMMARY_FOOTER_Y = 0.120
INSTAGRAM_SUMMARY_TITLE_SIZE = 24
INSTAGRAM_SUMMARY_SUBTITLE_SIZE = 16
INSTAGRAM_SUMMARY_STRIP_TEXT_SIZE = 20
INSTAGRAM_SUMMARY_SCORE_SIZE = 40
INSTAGRAM_SUMMARY_LABEL_SIZE = 13
INSTAGRAM_SUMMARY_VALUE_SIZE = 14
INSTAGRAM_SUMMARY_PARAGRAPH_SIZE = 12
INSTAGRAM_SUMMARY_LINE_SPACING = 1.17
INSTAGRAM_SUMMARY_STRIP_ROW_GAP = 0.030  # Gap between top result strip and first summary row
INSTAGRAM_SUMMARY_ROW_GAP = 0.017        # Gap/padding between summary rows
INSTAGRAM_SUMMARY_TOP_STRIP_LABEL = "EXPECTED PATTERN"  # Top green/orange/red strip text; keeps score badge from overlapping
INSTAGRAM_SUMMARY_DISCLAIMER_TEXT = "Disclaimer: Data-based estimate only. Not investment advice."
INSTAGRAM_SUMMARY_DISCLAIMER_SIZE = 9.0
INSTAGRAM_SUMMARY_DISCLAIMER_Y_SHIFT = -0.016
INSTAGRAM_SUMMARY_SHADOW_ALPHA = 0.10
INSTAGRAM_SUMMARY_SHOW_HANDLE = True
INSTAGRAM_SUMMARY_HANDLE_TEXT = "automationintrade"

# Summary - Standard
STANDARD_SUMMARY_FIG_W = 10.8
STANDARD_SUMMARY_FIG_H = 10.8
STANDARD_SUMMARY_DPI = 160
STANDARD_SUMMARY_CARD_X = 0.075
STANDARD_SUMMARY_CARD_Y = 0.145
STANDARD_SUMMARY_CARD_W = 0.85
STANDARD_SUMMARY_CARD_H = 0.66
STANDARD_SUMMARY_TITLE_Y = 0.890
STANDARD_SUMMARY_SUBTITLE_Y = 0.842
STANDARD_SUMMARY_FOOTER_Y = 0.090
STANDARD_SUMMARY_TITLE_SIZE = 28
STANDARD_SUMMARY_SUBTITLE_SIZE = 17
STANDARD_SUMMARY_STRIP_TEXT_SIZE = 21
STANDARD_SUMMARY_SCORE_SIZE = 46
STANDARD_SUMMARY_LABEL_SIZE = 14
STANDARD_SUMMARY_VALUE_SIZE = 16
STANDARD_SUMMARY_PARAGRAPH_SIZE = 14
STANDARD_SUMMARY_LINE_SPACING = 1.22
STANDARD_SUMMARY_STRIP_ROW_GAP = 0.032  # Gap between top result strip and first summary row
STANDARD_SUMMARY_ROW_GAP = 0.020        # Gap/padding between summary rows
STANDARD_SUMMARY_TOP_STRIP_LABEL = "EXPECTED PATTERN"  # Top green/orange/red strip text; keeps score badge from overlapping
STANDARD_SUMMARY_DISCLAIMER_TEXT = "Disclaimer: Data-based estimate only. Not investment advice."
STANDARD_SUMMARY_DISCLAIMER_SIZE = 9.5
STANDARD_SUMMARY_DISCLAIMER_Y_SHIFT = 0.000
STANDARD_SUMMARY_SHADOW_ALPHA = 0.12
STANDARD_SUMMARY_SHOW_HANDLE = True
STANDARD_SUMMARY_HANDLE_TEXT = "automationintrade"

# Summary - Reels
REELS_SUMMARY_FIG_W = 6.75
REELS_SUMMARY_FIG_H = 12.0
REELS_SUMMARY_DPI = 160
REELS_SUMMARY_CARD_X = 0.07
REELS_SUMMARY_CARD_Y = 0.18
REELS_SUMMARY_CARD_W = 0.86
REELS_SUMMARY_CARD_H = 0.62
REELS_SUMMARY_TITLE_Y = 0.895
REELS_SUMMARY_SUBTITLE_Y = 0.848
REELS_SUMMARY_FOOTER_Y = 0.115
REELS_SUMMARY_TITLE_SIZE = 23
REELS_SUMMARY_SUBTITLE_SIZE = 15
REELS_SUMMARY_STRIP_TEXT_SIZE = 18
REELS_SUMMARY_SCORE_SIZE = 38
REELS_SUMMARY_LABEL_SIZE = 12
REELS_SUMMARY_VALUE_SIZE = 13
REELS_SUMMARY_PARAGRAPH_SIZE = 11.5
REELS_SUMMARY_LINE_SPACING = 1.15
REELS_SUMMARY_STRIP_ROW_GAP = 0.028  # Gap between top result strip and first summary row
REELS_SUMMARY_ROW_GAP = 0.020        # Gap/padding between summary rows
REELS_SUMMARY_TOP_STRIP_LABEL = "EXPECTED PATTERN"  # Top green/orange/red strip text; keeps score badge from overlapping
REELS_SUMMARY_DISCLAIMER_TEXT = "Disclaimer: Data-based estimate only. Not investment advice."
REELS_SUMMARY_DISCLAIMER_SIZE = 9
REELS_SUMMARY_DISCLAIMER_Y_SHIFT = -0.015
REELS_SUMMARY_SHADOW_ALPHA = 0.11
REELS_SUMMARY_SHOW_HANDLE = True
REELS_SUMMARY_HANDLE_TEXT = "automationintrade"

# =========================================================
# GENERAL IMAGE BUTTONS
# =========================================================
GENERAL_FIG_W = 7.5
GENERAL_FIG_H = 12.0
GENERAL_DPI = 160

GENERAL_CARD_X = 0.06
GENERAL_CARD_Y = 0.14
GENERAL_CARD_W = 0.88
GENERAL_CARD_H = 0.70
GENERAL_STRIP_H = 0.065

GENERAL_TITLE_Y = 0.902
GENERAL_SUBTITLE_Y = 0.866
GENERAL_COMPANY_Y = 0.866
GENERAL_FOOTER_Y = 0.099

GENERAL_TITLE_SIZE = 26
GENERAL_SUBTITLE_SIZE = 16
GENERAL_COMPANY_SIZE = 16
GENERAL_STRIP_TEXT_SIZE = 20
GENERAL_BADGE_TEXT_SIZE = 16
GENERAL_TABLE_FONT_SIZE = 9.6
GENERAL_TABLE_SCALE_Y = 3.10
GENERAL_TABLE_TOP_GAP = 0.020
GENERAL_TABLE_BOTTOM_GAP = 0.010
GENERAL_TABLE_SIDE_PAD = 0.020
GENERAL_SHADOW_ALPHA = 0.12

GENERAL_SHOW_HANDLE = True
GENERAL_HANDLE_TEXT = "automationintrade"

# =========================================================
# INSTAGRAM IMAGE BUTTONS
# =========================================================
INSTAGRAM_FIG_W = 8.0
INSTAGRAM_FIG_H = 10.0
INSTAGRAM_DPI = 160

INSTAGRAM_CARD_X = 0.085
INSTAGRAM_CARD_Y = 0.16
INSTAGRAM_CARD_W = 0.83
INSTAGRAM_CARD_H = 0.626
INSTAGRAM_STRIP_H = 0.060

INSTAGRAM_TITLE_Y = 0.870
INSTAGRAM_SUBTITLE_Y = 0.820
INSTAGRAM_COMPANY_Y = 0.820
INSTAGRAM_FOOTER_Y = 0.120

INSTAGRAM_TITLE_SIZE = 24
INSTAGRAM_SUBTITLE_SIZE = 16
INSTAGRAM_COMPANY_SIZE = 16
INSTAGRAM_STRIP_TEXT_SIZE = 20
INSTAGRAM_BADGE_TEXT_SIZE = 14
INSTAGRAM_TABLE_FONT_SIZE = 10
INSTAGRAM_TABLE_SCALE_Y = 4.60
INSTAGRAM_TABLE_TOP_GAP = 0.014
INSTAGRAM_TABLE_BOTTOM_GAP = 0.012
INSTAGRAM_TABLE_SIDE_PAD = 0.018
INSTAGRAM_SHADOW_ALPHA = 0.10

INSTAGRAM_SHOW_HANDLE = True
INSTAGRAM_HANDLE_TEXT = "automationintrade"


def _clean_num(x):
    if x is None:
        return np.nan
    s = str(x).strip()
    s = s.replace(",", "").replace("—", "").strip()
    # keep minus if present
    if s == "":
        return np.nan
    try:
        return float(s)
    except:
        # try remove non-numeric (like %)
        s2 = "".join(ch for ch in s if (ch.isdigit() or ch in ".-"))
        try:
            return float(s2) if s2 else np.nan
        except:
            return np.nan


def parse_quarter_col_to_qend(col_label: str) -> pd.Timestamp:
    col_label = str(col_label).strip()
    dt = pd.to_datetime(col_label, format="%b %Y", errors="coerce")
    if pd.isna(dt):
        dt = pd.to_datetime(col_label, errors="coerce")
    if pd.isna(dt):
        raise ValueError(f"Cannot parse quarter label: {col_label}")
    return (dt + pd.offsets.MonthEnd(0)).normalize()




def safe_request_get(url: str, headers: dict | None = None, timeout: int = 30, max_retries: int | None = None):
    max_retries = SCREENER_MAX_RETRIES if max_retries is None else max_retries

    last_response = None
    for attempt in range(1, max_retries + 1):
        try:
            r = requests.get(url, headers=headers or {}, timeout=timeout)
            last_response = r
            if r.status_code == 429 and attempt < max_retries:
                print(f"Screener 429 rate-limit. Waiting {SCREENER_429_RETRY_SECONDS}s before retry {attempt + 1}/{max_retries}...")
                time.sleep(SCREENER_429_RETRY_SECONDS)
                continue
            return r
        except Exception as e:
            print(f"Request failed: {e}. Retry {attempt}/{max_retries}")
            if attempt < max_retries:
                time.sleep(2)
    return last_response


def fetch_quarterly_from_screener(url: str) -> pd.DataFrame:
    """
    Fetch ONLY the 'Quarterly Results' table from Screener.

    Fix for AUBANK:
    - Avoids accidentally picking old Annual/P&L table
    - First searches table near 'Quarterly Results' heading
    - Supports Sales or Revenue
    - Supports EPS / EPS in Rs
    - Returns: Quarter, Revenue, Expenses, Net Profit, EPS, QuarterEnd
    """
    import re
    import time
    from io import StringIO
    import pandas as pd
    import numpy as np
    from bs4 import BeautifulSoup

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://www.screener.in/",
    }

    quarter_pattern = re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}$", re.I)

    def _is_quarter_label(x) -> bool:
        return bool(quarter_pattern.match(str(x).strip()))

    def _norm_text(x) -> str:
        return str(x).replace("\xa0", " ").replace("+", "").strip()

    def _norm_metric(x) -> str:
        s = _norm_text(x).lower()
        s = re.sub(r"\s+", " ", s)
        return s

    def _fetch_html(u: str) -> str:
        time.sleep(SCREENER_REQUEST_SLEEP_SECONDS)

        r = safe_request_get(
            u,
            headers=headers,
            timeout=30,
            max_retries=SCREENER_MAX_RETRIES
        )

        if r is None or r.status_code != 200:
            raise RuntimeError(
                f"Failed to fetch Screener page. "
                f"Status={getattr(r, 'status_code', None)} URL={u}"
            )

        return r.text

    def _find_quarterly_table_html(html: str):
        soup = BeautifulSoup(html, "html.parser")

        # 1) Best method: find table after heading 'Quarterly Results'
        quarterly_heading = None
        for tag in soup.find_all(["h2", "h3", "h4", "p", "span"]):
            txt = _norm_text(tag.get_text(" ", strip=True)).lower()
            if "quarterly results" in txt:
                quarterly_heading = tag
                break

        if quarterly_heading:
            for table in quarterly_heading.find_all_next("table", limit=5):
                try:
                    dfs = pd.read_html(StringIO(str(table)))
                    if not dfs:
                        continue

                    tbl = dfs[0].copy()
                    flat = " | ".join(
                        _norm_text(v).lower()
                        for v in tbl.astype(str).fillna("").values.flatten()
                    )

                    quarter_count = len(re.findall(r"(mar|jun|sep|dec)\s+\d{4}", flat, flags=re.I))
                    has_sales = ("sales" in flat) or ("revenue" in flat)
                    has_np = "net profit" in flat
                    has_eps = "eps" in flat

                    if quarter_count >= 3 and has_sales and has_np and has_eps:
                        return tbl
                except Exception:
                    continue

        # 2) Fallback: score all tables, but prefer table with quarter labels
        try:
            tables = pd.read_html(StringIO(html))
        except Exception:
            return None

        best_table = None
        best_score = -1

        for tbl in tables:
            try:
                temp = tbl.copy()

                if isinstance(temp.columns, pd.MultiIndex):
                    temp.columns = [
                        " ".join([_norm_text(c) for c in col if _norm_text(c)]).strip()
                        for col in temp.columns
                    ]
                else:
                    temp.columns = [_norm_text(c) for c in temp.columns]

                flat = " | ".join(
                    _norm_text(v).lower()
                    for v in temp.astype(str).fillna("").values.flatten()
                )

                score = 0
                score += len(re.findall(r"(mar|jun|sep|dec)\s+\d{4}", flat, flags=re.I))

                for kw in ["sales", "revenue", "expenses", "net profit", "eps"]:
                    if kw in flat:
                        score += 3

                # Quarterly table should generally have many month-year columns
                if temp.shape[1] >= 5:
                    score += 2

                if score > best_score:
                    best_score = score
                    best_table = temp

            except Exception:
                continue

        return best_table if best_score >= 12 else None

    def _normalize_table(raw: pd.DataFrame) -> pd.DataFrame:
        df = raw.copy()

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [
                " ".join([_norm_text(c) for c in col if _norm_text(c)]).strip()
                for col in df.columns
            ]
        else:
            df.columns = [_norm_text(c) for c in df.columns]

        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all").reset_index(drop=True)

        for col in df.columns:
            df[col] = df[col].apply(_norm_text)

        # Sometimes first row contains actual quarter headers
        current_cols = [_norm_text(c) for c in df.columns.tolist()]
        first_row = [_norm_text(x) for x in df.iloc[0].tolist()] if not df.empty else []

        col_q_count = sum(_is_quarter_label(c) for c in current_cols)
        row_q_count = sum(_is_quarter_label(x) for x in first_row)

        if row_q_count >= 2 and row_q_count >= col_q_count:
            new_cols = []
            for i, val in enumerate(first_row):
                if i == 0:
                    new_cols.append("Metric")
                else:
                    new_cols.append(val)
            df = df.iloc[1:].reset_index(drop=True)
            df.columns = new_cols
        else:
            df.columns = ["Metric"] + current_cols[1:]

        quarter_cols = [c for c in df.columns if _is_quarter_label(c)]

        if len(quarter_cols) < 4:
            raise RuntimeError("Quarterly columns not detected correctly from Screener table.")

        df["MetricNorm"] = df["Metric"].apply(_norm_metric)

        def _get_row(possible_names):
            for name in possible_names:
                hit = df[df["MetricNorm"] == name]
                if not hit.empty:
                    return hit.iloc[0]

            # fallback contains match
            for name in possible_names:
                hit = df[df["MetricNorm"].str.contains(name, regex=False, na=False)]
                if not hit.empty:
                    return hit.iloc[0]

            return None

        revenue_row = _get_row(["sales", "revenue"])
        expenses_row = _get_row(["expenses"])
        net_profit_row = _get_row(["net profit"])
        eps_row = _get_row(["eps in rs", "eps"])

        if revenue_row is None or expenses_row is None or net_profit_row is None or eps_row is None:
            raise RuntimeError(
                "Required rows not found in Quarterly Results table. "
                "Needed: Sales/Revenue, Expenses, Net Profit, EPS."
            )

        out = pd.DataFrame({
            "Quarter": quarter_cols,
            "Revenue": [_clean_num(revenue_row[c]) for c in quarter_cols],
            "Expenses": [_clean_num(expenses_row[c]) for c in quarter_cols],
            "Net Profit": [_clean_num(net_profit_row[c]) for c in quarter_cols],
            "EPS": [_clean_num(eps_row[c]) for c in quarter_cols],
        })

        out = out.dropna(subset=["Revenue", "Expenses", "Net Profit", "EPS"]).copy()

        if out.empty:
            raise RuntimeError("Quarterly Results table became empty after numeric cleaning.")

        out["QuarterEnd"] = out["Quarter"].apply(parse_quarter_col_to_qend)
        out = out.sort_values("QuarterEnd").reset_index(drop=True)

        return out

    html = _fetch_html(url)
    raw_table = _find_quarterly_table_html(html)

    # If consolidated page fails, retry standalone
    if raw_table is None and "/consolidated" in url:
        fallback_url = url.replace("/consolidated/", "/")
        html = _fetch_html(fallback_url)
        raw_table = _find_quarterly_table_html(html)

    if raw_table is None:
        raise RuntimeError(
            "Could not locate the correct Quarterly Results table on Screener page."
        )

    result = _normalize_table(raw_table)

    print(
        f"✅ Parsed Screener quarterly table: "
        f"{result['Quarter'].iloc[0]} -> {result['Quarter'].iloc[-1]}"
    )

    return result


def build_features(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()

    d["Rev_QoQ%"] = d["Revenue"].pct_change(1) * 100
    d["NP_QoQ%"] = d["Net Profit"].pct_change(1) * 100
    d["EPS_QoQ%"] = d["EPS"].pct_change(1) * 100

    d["Rev_YoY%"] = d["Revenue"].pct_change(4) * 100
    d["NP_YoY%"] = d["Net Profit"].pct_change(4) * 100
    d["EPS_YoY%"] = d["EPS"].pct_change(4) * 100

    d["Exp_QoQ%"] = d["Expenses"].pct_change(1) * 100
    d["Exp_YoY%"] = d["Expenses"].pct_change(4) * 100

    d["NP_Margin%"] = (d["Net Profit"] / d["Revenue"]) * 100
    d["NP_Margin_Delta_QoQ"] = d["NP_Margin%"].diff(1)
    d["NP_Margin_Delta_YoY"] = d["NP_Margin%"].diff(4)

    d["SurpriseScoreRaw"] = (
        0.35 * d["NP_YoY%"].fillna(0)
        + 0.25 * d["EPS_YoY%"].fillna(0)
        + 0.20 * d["NP_QoQ%"].fillna(0)
        + 0.10 * d["NP_Margin_Delta_YoY"].fillna(0)
        + 0.10 * d["NP_Margin_Delta_QoQ"].fillna(0)
        - 0.10 * d["Exp_YoY%"].fillna(0)
        - 0.05 * d["Exp_QoQ%"].fillna(0)
    )

    roll = 8
    mu = d["SurpriseScoreRaw"].rolling(roll, min_periods=4).mean()
    sd = d["SurpriseScoreRaw"].rolling(roll, min_periods=4).std()
    d["SurpriseScore"] = (d["SurpriseScoreRaw"] - mu) / sd
    d["SurpriseScore"] = d["SurpriseScore"].replace([np.inf, -np.inf], np.nan).fillna(0)
    return d


def grade_from_marks(score: float) -> str:
    # 0-100
    if score < 35: return "Poor"
    if score < 55: return "Average"
    if score < 70: return "Good"
    if score < 85: return "Very Good"
    return "Excellent"


def clip01(x: float) -> float:
    return float(np.clip(x, 0.0, 1.0))

def _get_result_image_layout(instagram_mode: bool = False) -> dict:
    if instagram_mode:
        return {
            "figsize": (INSTAGRAM_FIG_W, INSTAGRAM_FIG_H),
            "dpi": INSTAGRAM_DPI,
            "card_x": INSTAGRAM_CARD_X,
            "card_y": INSTAGRAM_CARD_Y,
            "card_w": INSTAGRAM_CARD_W,
            "card_h": INSTAGRAM_CARD_H,
            "strip_h": INSTAGRAM_STRIP_H,
            "title_y": INSTAGRAM_TITLE_Y,
            "subtitle_y": INSTAGRAM_SUBTITLE_Y,
            "company_y": INSTAGRAM_COMPANY_Y,
            "footer_y": INSTAGRAM_FOOTER_Y,
            "title_size": INSTAGRAM_TITLE_SIZE,
            "subtitle_size": INSTAGRAM_SUBTITLE_SIZE,
            "company_size": INSTAGRAM_COMPANY_SIZE,
            "strip_text_size": INSTAGRAM_STRIP_TEXT_SIZE,
            "badge_text_size": INSTAGRAM_BADGE_TEXT_SIZE,
            "table_font_size": INSTAGRAM_TABLE_FONT_SIZE,
            "table_scale_y": INSTAGRAM_TABLE_SCALE_Y,
            "table_top_gap": INSTAGRAM_TABLE_TOP_GAP,
            "table_bottom_gap": INSTAGRAM_TABLE_BOTTOM_GAP,
            "table_side_pad": INSTAGRAM_TABLE_SIDE_PAD,
            "shadow_alpha": INSTAGRAM_SHADOW_ALPHA,
            "show_handle": INSTAGRAM_SHOW_HANDLE,
            "handle_text": INSTAGRAM_HANDLE_TEXT,
        }

    return {
        "figsize": (GENERAL_FIG_W, GENERAL_FIG_H),
        "dpi": GENERAL_DPI,
        "card_x": GENERAL_CARD_X,
        "card_y": GENERAL_CARD_Y,
        "card_w": GENERAL_CARD_W,
        "card_h": GENERAL_CARD_H,
        "strip_h": GENERAL_STRIP_H,
        "title_y": GENERAL_TITLE_Y,
        "subtitle_y": GENERAL_SUBTITLE_Y,
        "company_y": GENERAL_COMPANY_Y,
        "footer_y": GENERAL_FOOTER_Y,
        "title_size": GENERAL_TITLE_SIZE,
        "subtitle_size": GENERAL_SUBTITLE_SIZE,
        "company_size": GENERAL_COMPANY_SIZE,
        "strip_text_size": GENERAL_STRIP_TEXT_SIZE,
        "badge_text_size": GENERAL_BADGE_TEXT_SIZE,
        "table_font_size": GENERAL_TABLE_FONT_SIZE,
        "table_scale_y": GENERAL_TABLE_SCALE_Y,
        "table_top_gap": GENERAL_TABLE_TOP_GAP,
        "table_bottom_gap": GENERAL_TABLE_BOTTOM_GAP,
        "table_side_pad": GENERAL_TABLE_SIDE_PAD,
        "shadow_alpha": GENERAL_SHADOW_ALPHA,
        "show_handle": GENERAL_SHOW_HANDLE,
        "handle_text": GENERAL_HANDLE_TEXT,
    }


def export_result_table_as_image(
    symbol: str,
    company_name: str,
    score_df: pd.DataFrame,
    total_row: dict | None = None,
    out_path_template: str = "Image/{SYMBOL}.jpeg",
    title_suffix: str = "Q4 Results 2026",
    footer_text: str = "",
    instagram_mode: bool = False
) -> str:
    import os
    import pandas as pd
    import matplotlib.pyplot as plt
    from matplotlib.patches import FancyBboxPatch

    out_path = out_path_template.replace("{SYMBOL}", symbol)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    df = score_df.copy()
    if total_row is not None:
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
    df = df.fillna("")
    cfg = _get_result_image_layout(instagram_mode)

    BG_NAVY = "#0B1E3A"
    GOLD_TITLE = "#E5C26A"
    TEXT_LIGHT = "#E8EDF4"
    CARD_BG = "#E9EDF2"
    GRID_BORDER = "#CFCFCF"
    HEADER_BG = "#DCE2EA"
    TOTAL_BG = "#FFF2CC"

    STRIP_EXCELLENT = "#1F6F4A"
    STRIP_STRONG = "#2E8B57"
    STRIP_STABLE = "#46A36A"
    STRIP_SOFT = "#E39D2C"
    STRIP_POOR = "#C43C3C"

    total_pct = None
    grade_val = ""
    verdict_text = "SOFT"
    strip_color = STRIP_SOFT
    score_text = ""

    try:
        trow = df.loc[df["Parameter"].astype(str).str.upper() == "TOTAL"]
        if not trow.empty:
            grade_val = str(trow.iloc[0].get("Grade", "")).strip()
            tp = trow.iloc[0].get("Total%", "")
            total_pct = float(tp) if tp != "" else None
    except Exception:
        pass

    grade_norm = (grade_val or "").lower().strip()
    if grade_norm == "excellent":
        verdict_text = "EXCELLENT"
        strip_color = STRIP_EXCELLENT
    elif grade_norm == "very good":
        verdict_text = "STRONG"
        strip_color = STRIP_STRONG
    elif grade_norm == "good":
        verdict_text = "STABLE"
        strip_color = STRIP_STABLE
    elif grade_norm == "average":
        verdict_text = "SOFT"
        strip_color = STRIP_SOFT
    elif grade_norm == "poor":
        verdict_text = "POOR"
        strip_color = STRIP_POOR
    else:
        verdict_text = grade_val.upper() if grade_val else "SOFT"
        strip_color = STRIP_SOFT

    score_text = f"{total_pct:.1f}%" if isinstance(total_pct, (int, float)) else ""
    orig_cols = df.columns.tolist()

    def _pretty_col(c: str) -> str:
        c = str(c).strip()
        m = {
            "MaxMarks": "Max\nMarks",
            "Weightage%": "Weight\nage",
            "Score(0-1)": "Score\n(0-1)",
            "Total%": "Total\n%",
        }
        return m.get(c, c)

    col_labels_display = [_pretty_col(c) for c in orig_cols]

    fig = plt.figure(figsize=cfg["figsize"], dpi=cfg["dpi"])
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_axis_off()
    fig.patch.set_facecolor(BG_NAVY)

    def _text_width_px(text, fontsize, fontweight="bold"):
        """
        Measure rendered text width in pixels using matplotlib renderer.
        """
        temp = ax.text(
            0, 0, text,
            fontsize=fontsize,
            fontweight=fontweight,
            alpha=0
        )
        fig.canvas.draw()
        renderer = fig.canvas.get_renderer()
        width = temp.get_window_extent(renderer=renderer).width
        temp.remove()
        return width


    def _fit_company_title_to_card_width(
        text,
        max_width_px,
        fontsize,
        fontweight="bold",
        max_lines=2
    ):
        """
        Logic:
        1. If title fits, keep as-is.
        2. If adding a full word overflows badly, move that word to next line.
        3. If overflow is only around 1-2 letters, trim those extra letters.
        4. If still too long after max_lines, trim last line safely.
        """
        text = "" if text is None else str(text).strip()
        words = text.split()

        if not words:
            return ""

        # Already fits
        if _text_width_px(text, fontsize, fontweight) <= max_width_px:
            return text

        lines = []
        current = ""

        # Approx width of 2 letters. Used to decide small overflow trimming.
        two_letter_width = _text_width_px("WW", fontsize, fontweight)

        for word in words:
            candidate = word if not current else current + " " + word
            candidate_width = _text_width_px(candidate, fontsize, fontweight)

            if candidate_width <= max_width_px:
                current = candidate
                continue

            overflow_px = candidate_width - max_width_px

            # If overflow is very small, trim 1-2 trailing characters
            if overflow_px <= two_letter_width:
                trimmed = candidate
                while trimmed and _text_width_px(trimmed, fontsize, fontweight) > max_width_px:
                    trimmed = trimmed[:-1].rstrip()
                current = trimmed
                continue

            # Otherwise move full overflowing word to next line
            if current:
                lines.append(current)
                current = word
            else:
                # Single very long word: trim it safely
                trimmed = word
                while trimmed and _text_width_px(trimmed, fontsize, fontweight) > max_width_px:
                    trimmed = trimmed[:-1].rstrip()
                current = trimmed

            if len(lines) >= max_lines:
                break

        if current and len(lines) < max_lines:
            lines.append(current)

        # Safety: trim last line if still overflowing
        if lines:
            last = lines[-1]
            while last and _text_width_px(last, fontsize, fontweight) > max_width_px:
                last = last[:-1].rstrip()
            lines[-1] = last

        return "\n".join(lines[:max_lines])
    
    card_x, card_y, card_w, card_h = cfg["card_x"], cfg["card_y"], cfg["card_w"], cfg["card_h"]

    max_title_width_px = fig.bbox.width * card_w * 0.98

    fitted_company_name = _fit_company_title_to_card_width(
        text=company_name,
        max_width_px=max_title_width_px,
        fontsize=cfg["title_size"],
        fontweight="bold",
        max_lines=2
    )

    title_line_count = fitted_company_name.count("\n") + 1

    # ---------------------------------------------------------
    # AUTO TITLE POSITION FIX
    # ---------------------------------------------------------
    # When company name becomes 2 lines, move ONLY the title upward.
    # Keep subtitle mostly in its original place so the layout remains stable.
    if title_line_count > 1:
        title_y = cfg["title_y"] + (0.022 if instagram_mode else 0.018)
        subtitle_y = cfg["company_y"] - (0.004 if instagram_mode else 0.002)
    else:
        title_y = cfg["title_y"]
        subtitle_y = cfg["company_y"]

    ax.text(
        0.5, title_y, fitted_company_name,
        ha="center", va="center",
        fontsize=cfg["title_size"],
        fontweight="bold",
        color=GOLD_TITLE,
        linespacing=0.92 if title_line_count > 1 else 1.05
    )

    ax.text(
        0.5, subtitle_y, f"{title_suffix}",
        ha="center", va="center",
        fontsize=cfg["company_size"],
        color=TEXT_LIGHT,
        alpha=0.9
    )
    #card_x, card_y, card_w, card_h = cfg["card_x"], cfg["card_y"], cfg["card_w"], cfg["card_h"]

    shadow = FancyBboxPatch(
        (card_x, card_y - 0.006), card_w, card_h,
        boxstyle="round,pad=0.012,rounding_size=0.03",
        linewidth=0, facecolor="black", alpha=cfg["shadow_alpha"],
        transform=ax.transAxes, zorder=0
    )
    ax.add_patch(shadow)

    card = FancyBboxPatch(
        (card_x, card_y), card_w, card_h,
        boxstyle="round,pad=0.012,rounding_size=0.03",
        linewidth=1.5, edgecolor="#D6D9DE", facecolor=CARD_BG,
        transform=ax.transAxes, zorder=1
    )
    ax.add_patch(card)

    strip_h = cfg["strip_h"]
    strip_y = card_y + card_h - strip_h - cfg["table_top_gap"]

    strip = FancyBboxPatch(
        (card_x + 0.02, strip_y), card_w - 0.04, strip_h,
        boxstyle="round,pad=0.010,rounding_size=0.02",
        linewidth=0, facecolor=strip_color,
        transform=ax.transAxes, zorder=3
    )
    ax.add_patch(strip)

    ax.text(
        card_x + 0.045, strip_y + strip_h / 2,
        f"OVERALL RESULT: {verdict_text}",
        transform=ax.transAxes, ha="left", va="center",
        fontsize=cfg["strip_text_size"], fontweight="bold",
        color="white", zorder=4
    )

    badge_w, badge_h = 0.23, strip_h * 0.78
    badge_x = card_x + card_w - badge_w - 0.035
    badge_y = strip_y + (strip_h - badge_h) / 2

    badge = FancyBboxPatch(
        (badge_x, badge_y), badge_w, badge_h,
        boxstyle="round,pad=0.008,rounding_size=0.03",
        linewidth=0, facecolor="#FFFFFF", alpha=0.16,
        transform=ax.transAxes, zorder=5
    )
    ax.add_patch(badge)

    ax.text(
        badge_x + badge_w / 2, badge_y + badge_h / 2,
        f"Score: {score_text}",
        transform=ax.transAxes, ha="center", va="center",
        fontsize=cfg["badge_text_size"], fontweight="bold",
        color="white", zorder=6
    )

    table_bbox = [
        card_x + cfg["table_side_pad"],
        card_y + cfg["table_bottom_gap"],
        card_w - (2 * cfg["table_side_pad"]),
        card_h - strip_h - cfg["table_bottom_gap"] - cfg["table_top_gap"] - 0.03,
    ]

    table = ax.table(
        cellText=df.values,
        colLabels=col_labels_display,
        cellLoc="left",
        colLoc="center",
        loc="upper left",
        bbox=table_bbox
    )
    table.set_zorder(10)
    for cell in table.get_celld().values():
        cell.set_zorder(10)

    table.auto_set_font_size(False)
    table.set_fontsize(cfg["table_font_size"])
    table.scale(1, cfg["table_scale_y"])

    col_widths = {
        0: 0.255, 1: 0.095, 2: 0.105, 3: 0.125,
        4: 0.105, 5: 0.085, 6: 0.085, 7: 0.105
    }
    for (r, c), cell in table.get_celld().items():
        if c in col_widths:
            cell.set_width(col_widths[c])

    ncols = len(orig_cols)
    for c in range(ncols):
        hcell = table[(0, c)]
        hcell.set_height(hcell.get_height() * (1.65 if instagram_mode else 1.75))
        hcell.PAD = 0.25
        htxt = hcell.get_text()
        htxt.set_va("center")
        htxt.set_ha("center")
        htxt.set_wrap(True)

    grade_col_idx = None
    for i, name in enumerate(orig_cols):
        if str(name).strip().lower() == "grade":
            grade_col_idx = i
            break

    if grade_col_idx is not None:
        for r in range(1, len(df) + 1):
            cell = table[(r, grade_col_idx)]
            txt_obj = cell.get_text()
            raw = str(txt_obj.get_text()).strip()
            if " " in raw and len(raw) >= 8:
                parts = raw.replace("  ", " ").strip().split(" ")
                if len(parts) >= 2:
                    txt_obj.set_text(parts[0] + "\n" + " ".join(parts[1:]))
            txt_obj.set_wrap(True)
            txt_obj.set_va("center")
            txt_obj.set_ha("center")
            cell.PAD = 0.18

    for (r, c), cell in table.get_celld().items():
        cell.set_edgecolor(GRID_BORDER)
        cell.set_linewidth(0.8)
        if r == 0:
            cell.set_facecolor(HEADER_BG)
            cell.set_text_props(weight="bold")
        else:
            is_total = str(df.iloc[r - 1, 0]).upper().strip() == "TOTAL"
            if is_total:
                cell.set_facecolor(TOTAL_BG)
                cell.set_text_props(weight="bold")
            else:
                cell.set_facecolor(CARD_BG)

    # Footer handle below box
    if cfg.get("show_handle", False):
        fig.text(
            0.5, cfg["footer_y"], cfg.get("handle_text", "automationintrade"),
            ha="center", va="center",
            fontsize=12, color=TEXT_LIGHT, fontweight="bold"
        )

    # Optional separate footer text, only if you want extra text later
    if footer_text:
        fig.text(
            0.5, max(cfg["footer_y"] - 0.022, 0.01), footer_text,
            ha="center", va="center",
            fontsize=10, color=TEXT_LIGHT
        )

    plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
    fig.savefig(out_path, bbox_inches="tight", pad_inches=0)
    plt.close(fig)
    return out_path

def score_quarter_parameters(fdf: pd.DataFrame) -> pd.DataFrame:
    """
    Builds 20-parameter scoring for the LATEST quarter using:
    - YoY / QoQ growth
    - Margin changes
    - Expense control
    - Trend persistence (3Q)
    - Stability (rolling std)
    All scores normalized 0..1 then converted to marks.
    """
    d = fdf.copy()
    latest = d.iloc[-1].copy()

    # ===== helper metrics (safe defaults) =====
    np_yoy = float(latest.get("NP_YoY%", 0) if pd.notna(latest.get("NP_YoY%", np.nan)) else 0)
    eps_yoy = float(latest.get("EPS_YoY%", 0) if pd.notna(latest.get("EPS_YoY%", np.nan)) else 0)
    rev_yoy = float(latest.get("Rev_YoY%", 0) if pd.notna(latest.get("Rev_YoY%", np.nan)) else 0)

    np_qoq = float(latest.get("NP_QoQ%", 0) if pd.notna(latest.get("NP_QoQ%", np.nan)) else 0)
    eps_qoq = float(latest.get("EPS_QoQ%", 0) if pd.notna(latest.get("EPS_QoQ%", np.nan)) else 0)
    rev_qoq = float(latest.get("Rev_QoQ%", 0) if pd.notna(latest.get("Rev_QoQ%", np.nan)) else 0)

    exp_yoy = float(latest.get("Exp_YoY%", 0) if pd.notna(latest.get("Exp_YoY%", np.nan)) else 0)
    exp_qoq = float(latest.get("Exp_QoQ%", 0) if pd.notna(latest.get("Exp_QoQ%", np.nan)) else 0)

    margin = float(latest.get("NP_Margin%", 0) if pd.notna(latest.get("NP_Margin%", np.nan)) else 0)
    margin_dyoy = float(latest.get("NP_Margin_Delta_YoY", 0) if pd.notna(latest.get("NP_Margin_Delta_YoY", np.nan)) else 0)
    margin_dqoq = float(latest.get("NP_Margin_Delta_QoQ", 0) if pd.notna(latest.get("NP_Margin_Delta_QoQ", np.nan)) else 0)

    # trend persistence (last 3 quarters)
    def last_n_mean(col, n=3):
        s = d[col].tail(n)
        s = s.replace([np.inf, -np.inf], np.nan).dropna()
        return float(s.mean()) if len(s) else 0.0

    np_yoy_3q = last_n_mean("NP_YoY%", 3)
    eps_yoy_3q = last_n_mean("EPS_YoY%", 3)
    rev_yoy_3q = last_n_mean("Rev_YoY%", 3)

    # stability: rolling std of QoQ for revenue/NP (lower std = better)
    def roll_std(col, win=6):
        s = d[col].tail(win).replace([np.inf, -np.inf], np.nan).dropna()
        return float(s.std()) if len(s) else 0.0

    rev_qoq_std = roll_std("Rev_QoQ%", 6)
    np_qoq_std = roll_std("NP_QoQ%", 6)

    # SurpriseScore already z-normalized around 0
    surprise = float(latest.get("SurpriseScore", 0) if pd.notna(latest.get("SurpriseScore", np.nan)) else 0)

    # ===== scoring functions (0..1) =====
    # For growth: map -20%..+40% into 0..1
    def score_growth(pct, lo=-20, hi=40):
        return clip01((pct - lo) / (hi - lo))

    # For expense growth: lower is better. Map -10..+25 so -10 => 1, +25 => 0
    def score_expense(pct, lo=-10, hi=25):
        return clip01((hi - pct) / (hi - lo))

    # For margin level: map 5%..35% to 0..1
    def score_margin_level(m, lo=5, hi=35):
        return clip01((m - lo) / (hi - lo))

    # For margin change: map -2pp..+2pp to 0..1
    def score_margin_delta(pp, lo=-2.0, hi=2.0):
        return clip01((pp - lo) / (hi - lo))

    # For stability: lower std better. Map 0..15 so 0=>1, 15=>0
    def score_stability(std, hi=15.0):
        return clip01((hi - std) / hi)

    # SurpriseScore: map -1.5..+1.5 to 0..1
    def score_surprise(z, lo=-1.5, hi=1.5):
        return clip01((z - lo) / (hi - lo))

    # ===== Define 20 parameters =====
    # priority: High/Medium/Low (your reporting)
    params = [
        # Profitability & earnings
        ("NetProfit YoY%", "High", 8, 8, score_growth(np_yoy)),
        ("EPS YoY%", "High", 7, 7, score_growth(eps_yoy)),
        ("NetProfit QoQ%", "High", 6, 6, score_growth(np_qoq, lo=-15, hi=25)),
        ("EPS QoQ%", "High", 5, 5, score_growth(eps_qoq, lo=-15, hi=25)),

        # Growth quality
        ("Revenue YoY%", "High", 6, 6, score_growth(rev_yoy, lo=-10, hi=25)),
        ("Revenue QoQ%", "Medium", 4, 4, score_growth(rev_qoq, lo=-8, hi=12)),

        # Margin quality
        ("NP Margin%", "High", 6, 6, score_margin_level(margin)),
        ("Margin ΔYoY (pp)", "High", 5, 5, score_margin_delta(margin_dyoy)),
        ("Margin ΔQoQ (pp)", "Medium", 4, 4, score_margin_delta(margin_dqoq)),

        # Cost control
        ("Expenses YoY%", "High", 6, 6, score_expense(exp_yoy)),
        ("Expenses QoQ%", "Medium", 4, 4, score_expense(exp_qoq, lo=-5, hi=15)),

        # Consistency (last 3 quarters trend)
        ("NP YoY 3Q Avg%", "High", 5, 5, score_growth(np_yoy_3q)),
        ("EPS YoY 3Q Avg%", "High", 4, 4, score_growth(eps_yoy_3q)),
        ("Revenue YoY 3Q Avg%", "Medium", 4, 4, score_growth(rev_yoy_3q, lo=-5, hi=18)),

        # Stability
        ("Revenue QoQ Std(6Q)", "Medium", 4, 4, score_stability(rev_qoq_std)),
        ("NP QoQ Std(6Q)", "Medium", 4, 4, score_stability(np_qoq_std)),

        # Composite / market reaction proxy
        ("SurpriseScore", "High", 8, 8, score_surprise(surprise)),

        # Two extra “quality checks” to reach 20
        # 1) Profitability relative to costs: NP growth - Exp growth (YoY)
        ("Profit vs Cost Spread YoY", "High", 6, 6, score_growth(np_yoy - exp_yoy, lo=-20, hi=40)),
        # 2) Earnings acceleration: (QoQ NP) - (prev QoQ NP) if available
        ("NP Acceleration QoQ", "Low", 3, 3, score_growth(np_qoq - float(d["NP_QoQ%"].iloc[-2] if len(d) > 1 and pd.notna(d["NP_QoQ%"].iloc[-2]) else 0), lo=-20, hi=20)),
    ]

    # total max marks
    total_max = sum(p[2] for p in params)

    rows = []
    for name, priority, max_marks, weight, s01 in params:
        marks = round(max_marks * float(s01), 2)
        weight_pct = round((max_marks / total_max) * 100, 2)  # derived weightage %
        rows.append({
            "Parameter": name,
            "Priority": priority,
            "MaxMarks": max_marks,
            "Weightage%": weight_pct,
            "Score(0-1)": round(float(s01), 4),
            "Marks": marks
        })

    score_df = pd.DataFrame(rows)
    total_marks = round(float(score_df["Marks"].sum()), 2)
    total_pct = round((total_marks / total_max) * 100, 2)
    grade = grade_from_marks(total_pct)

    # Add summary row
    summary_row = pd.DataFrame([{
        "Parameter": "TOTAL",
        "Priority": "",
        "MaxMarks": total_max,
        "Weightage%": 100.0,
        "Score(0-1)": round(total_pct / 100.0, 4),
        "Marks": total_marks
    }])

    score_df = pd.concat([score_df, summary_row], ignore_index=True)
    score_df["Total%"] = ""
    score_df["Grade"] = ""
    score_df.loc[score_df["Parameter"] == "TOTAL", "Total%"] = total_pct
    score_df.loc[score_df["Parameter"] == "TOTAL", "Grade"] = grade

    return score_df


def title_case_stock_name(name: str) -> str:
    s = "" if name is None else str(name).strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return ""
    return " ".join(w[:1].upper() + w[1:].lower() if w else "" for w in s.split())


def parse_user_date_to_dd_mm_yyyy(date_text: str) -> str:
    dt = pd.to_datetime(str(date_text).strip(), dayfirst=True, errors="coerce")
    if pd.isna(dt):
        raise ValueError(f"Invalid DATE format: {date_text}. Use DD/MM/YYYY")
    return dt.strftime("%d_%m_%Y")


def resolve_result_calendar_stocklist_path(date_text: str, stock_list_path: str) -> str:
    dd_mm_yyyy = parse_user_date_to_dd_mm_yyyy(date_text)
    return str(stock_list_path).replace("{DD_MM_YYYY}", dd_mm_yyyy)


def build_screener_url_from_symbol(symbol: str) -> str:
    clean_symbol = str(symbol).strip().upper().replace(".NS", "").replace(".BO", "")
    return f"https://www.screener.in/company/{clean_symbol}/consolidated/"



def _clean_company_name_from_web(raw_name: str, fallback_symbol: str = "") -> str:
    """
    Clean a company name fetched from Screener/Yahoo style text.
    Keeps useful legal suffixes like Ltd, Limited, Bank, etc.
    """
    name = "" if raw_name is None else str(raw_name).strip()
    name = re.sub(r"\s+", " ", name).strip()

    # Remove common title fragments if the value came from <title> instead of <h1>.
    cleanup_patterns = [
        r"\s*[-|:]\s*Share Price.*$",
        r"\s*[-|:]\s*Financial.*$",
        r"\s*[-|:]\s*Screener.*$",
        r"\s*[-|:]\s*Stock.*$",
    ]
    for pat in cleanup_patterns:
        name = re.sub(pat, "", name, flags=re.I).strip()

    # Avoid returning the raw symbol as a display name.
    fallback_clean = str(fallback_symbol or "").strip().upper().replace(".NS", "").replace(".BO", "")
    if not name or name.lower() in ("nan", "none", "null") or name.upper() == fallback_clean:
        return ""

    return title_case_stock_name(name)


def fetch_company_name_from_screener(screener_url: str, fallback_symbol: str = "") -> str:
    """
    Used only when DATE is blank and the user runs:
        python .\ResultImpactPredictor_v6_fixed.py --symbol BAJAJHLDNG

    In DATE-based mode, the existing Excel Stock column still remains the source of truth.
    """
    try:
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": "https://www.screener.in/",
        }
        r = safe_request_get(screener_url, headers=headers, timeout=30, max_retries=SCREENER_MAX_RETRIES)
        if r is None or r.status_code != 200:
            return ""

        soup = BeautifulSoup(r.text, "html.parser")

        # Screener normally keeps the display company name in h1.
        for tag in soup.find_all(["h1", "h2"], limit=5):
            candidate = _clean_company_name_from_web(tag.get_text(" ", strip=True), fallback_symbol)
            if candidate:
                return candidate

        # Fallback to page title if h1 is not available.
        if soup.title:
            candidate = _clean_company_name_from_web(soup.title.get_text(" ", strip=True), fallback_symbol)
            if candidate:
                return candidate

    except Exception as e:
        print(f"⚠️ Could not fetch company name from Screener for {fallback_symbol}: {e}")

    return ""


def read_result_calendar_stocklist_excel(date_text: str, stock_list_path: str) -> pd.DataFrame:
    final_path = resolve_result_calendar_stocklist_path(date_text, stock_list_path)

    if not os.path.exists(final_path):
        raise FileNotFoundError(
            f"Result calendar StockList file not found: {final_path}\n"
            f"Check DATE='{date_text}' and file name ResultCalendar_{parse_user_date_to_dd_mm_yyyy(date_text)}.xlsx"
        )

    df = pd.read_excel(final_path)
    df.columns = [str(c).strip() for c in df.columns]

    required_cols = {"Stock", "Symbol"}
    missing = required_cols - set(df.columns)
    if missing:
        raise ValueError(f"StockList Excel must contain columns Stock and Symbol. Missing: {sorted(missing)}")

    df = df[["Stock", "Symbol"]].copy()
    df["Stock"] = df["Stock"].astype(str).str.strip()
    df["Symbol"] = df["Symbol"].astype(str).str.strip().str.upper()
    df["Symbol"] = df["Symbol"].replace({"NAN": "", "NONE": "", "NULL": ""})

    df = df[(df["Stock"] != "") & (~df["Stock"].str.lower().isin(["nan", "none", "null"]))].copy()
    df["CompanyName"] = df["Stock"].map(title_case_stock_name)

    blank_symbols = df[df["Symbol"] == ""].copy()
    if not blank_symbols.empty:
        os.makedirs("Result", exist_ok=True)
        skipped_file = f"Result/Skipped_Blank_Symbols_{parse_user_date_to_dd_mm_yyyy(date_text)}.csv"
        blank_symbols.to_csv(skipped_file, index=False)
        print(f"⚠️ Skipping {len(blank_symbols)} rows with blank/nan Symbol. Saved: {skipped_file}")

    if SKIP_ROWS_WITH_BLANK_SYMBOL:
        df = df[df["Symbol"] != ""].copy()

    df = df.drop_duplicates(subset=["Symbol"]).reset_index(drop=True)
    if df.empty:
        raise RuntimeError(f"No valid rows with Symbol found in StockList Excel: {final_path}")

    df["ScreenerURL"] = df["Symbol"].map(build_screener_url_from_symbol)

    print(f"Loaded {len(df)} stocks from local ResultCalendar file only: {final_path}")
    print(df[["CompanyName", "Symbol"]].to_string(index=False))
    return df




def _summary_grade_to_verdict_and_color(grade_val: str) -> tuple[str, str]:
    grade_norm = str(grade_val or "").lower().strip()
    if grade_norm == "excellent":
        return "EXCELLENT", "#1F6F4A"
    if grade_norm == "very good":
        return "STRONG", "#2E8B57"
    if grade_norm == "good":
        return "STABLE", "#46A36A"
    if grade_norm == "average":
        return "SOFT", "#E39D2C"
    if grade_norm == "poor":
        return "POOR", "#C43C3C"
    return (str(grade_val).upper().strip() if str(grade_val).strip() else "SOFT"), "#E39D2C"


def _get_summary_image_layout(image_type: str = "general") -> dict:
    image_type = str(image_type or "general").strip().lower()

    if image_type == "instagram":
        return {
            "figsize": (INSTAGRAM_SUMMARY_FIG_W, INSTAGRAM_SUMMARY_FIG_H),
            "dpi": INSTAGRAM_SUMMARY_DPI,
            "card_x": INSTAGRAM_SUMMARY_CARD_X,
            "card_y": INSTAGRAM_SUMMARY_CARD_Y,
            "card_w": INSTAGRAM_SUMMARY_CARD_W,
            "card_h": INSTAGRAM_SUMMARY_CARD_H,
            "title_y": INSTAGRAM_SUMMARY_TITLE_Y,
            "subtitle_y": INSTAGRAM_SUMMARY_SUBTITLE_Y,
            "footer_y": INSTAGRAM_SUMMARY_FOOTER_Y,
            "title_size": INSTAGRAM_SUMMARY_TITLE_SIZE,
            "subtitle_size": INSTAGRAM_SUMMARY_SUBTITLE_SIZE,
            "strip_text_size": INSTAGRAM_SUMMARY_STRIP_TEXT_SIZE,
            "score_size": INSTAGRAM_SUMMARY_SCORE_SIZE,
            "label_size": INSTAGRAM_SUMMARY_LABEL_SIZE,
            "value_size": INSTAGRAM_SUMMARY_VALUE_SIZE,
            "paragraph_size": INSTAGRAM_SUMMARY_PARAGRAPH_SIZE,
            "line_spacing": INSTAGRAM_SUMMARY_LINE_SPACING,
            "strip_row_gap": INSTAGRAM_SUMMARY_STRIP_ROW_GAP,
            "row_gap": INSTAGRAM_SUMMARY_ROW_GAP,
            "top_strip_label": INSTAGRAM_SUMMARY_TOP_STRIP_LABEL,
            "disclaimer_text": INSTAGRAM_SUMMARY_DISCLAIMER_TEXT,
            "disclaimer_size": INSTAGRAM_SUMMARY_DISCLAIMER_SIZE,
            "disclaimer_y_shift": INSTAGRAM_SUMMARY_DISCLAIMER_Y_SHIFT,
            "shadow_alpha": INSTAGRAM_SUMMARY_SHADOW_ALPHA,
            "show_handle": INSTAGRAM_SUMMARY_SHOW_HANDLE,
            "handle_text": INSTAGRAM_SUMMARY_HANDLE_TEXT,
        }

    if image_type == "standard":
        return {
            "figsize": (STANDARD_SUMMARY_FIG_W, STANDARD_SUMMARY_FIG_H),
            "dpi": STANDARD_SUMMARY_DPI,
            "card_x": STANDARD_SUMMARY_CARD_X,
            "card_y": STANDARD_SUMMARY_CARD_Y,
            "card_w": STANDARD_SUMMARY_CARD_W,
            "card_h": STANDARD_SUMMARY_CARD_H,
            "title_y": STANDARD_SUMMARY_TITLE_Y,
            "subtitle_y": STANDARD_SUMMARY_SUBTITLE_Y,
            "footer_y": STANDARD_SUMMARY_FOOTER_Y,
            "title_size": STANDARD_SUMMARY_TITLE_SIZE,
            "subtitle_size": STANDARD_SUMMARY_SUBTITLE_SIZE,
            "strip_text_size": STANDARD_SUMMARY_STRIP_TEXT_SIZE,
            "score_size": STANDARD_SUMMARY_SCORE_SIZE,
            "label_size": STANDARD_SUMMARY_LABEL_SIZE,
            "value_size": STANDARD_SUMMARY_VALUE_SIZE,
            "paragraph_size": STANDARD_SUMMARY_PARAGRAPH_SIZE,
            "line_spacing": STANDARD_SUMMARY_LINE_SPACING,
            "strip_row_gap": STANDARD_SUMMARY_STRIP_ROW_GAP,
            "row_gap": STANDARD_SUMMARY_ROW_GAP,
            "top_strip_label": STANDARD_SUMMARY_TOP_STRIP_LABEL,
            "disclaimer_text": STANDARD_SUMMARY_DISCLAIMER_TEXT,
            "disclaimer_size": STANDARD_SUMMARY_DISCLAIMER_SIZE,
            "disclaimer_y_shift": STANDARD_SUMMARY_DISCLAIMER_Y_SHIFT,
            "shadow_alpha": STANDARD_SUMMARY_SHADOW_ALPHA,
            "show_handle": STANDARD_SUMMARY_SHOW_HANDLE,
            "handle_text": STANDARD_SUMMARY_HANDLE_TEXT,
        }

    if image_type == "reels":
        return {
            "figsize": (REELS_SUMMARY_FIG_W, REELS_SUMMARY_FIG_H),
            "dpi": REELS_SUMMARY_DPI,
            "card_x": REELS_SUMMARY_CARD_X,
            "card_y": REELS_SUMMARY_CARD_Y,
            "card_w": REELS_SUMMARY_CARD_W,
            "card_h": REELS_SUMMARY_CARD_H,
            "title_y": REELS_SUMMARY_TITLE_Y,
            "subtitle_y": REELS_SUMMARY_SUBTITLE_Y,
            "footer_y": REELS_SUMMARY_FOOTER_Y,
            "title_size": REELS_SUMMARY_TITLE_SIZE,
            "subtitle_size": REELS_SUMMARY_SUBTITLE_SIZE,
            "strip_text_size": REELS_SUMMARY_STRIP_TEXT_SIZE,
            "score_size": REELS_SUMMARY_SCORE_SIZE,
            "label_size": REELS_SUMMARY_LABEL_SIZE,
            "value_size": REELS_SUMMARY_VALUE_SIZE,
            "paragraph_size": REELS_SUMMARY_PARAGRAPH_SIZE,
            "line_spacing": REELS_SUMMARY_LINE_SPACING,
            "strip_row_gap": REELS_SUMMARY_STRIP_ROW_GAP,
            "row_gap": REELS_SUMMARY_ROW_GAP,
            "top_strip_label": REELS_SUMMARY_TOP_STRIP_LABEL,
            "disclaimer_text": REELS_SUMMARY_DISCLAIMER_TEXT,
            "disclaimer_size": REELS_SUMMARY_DISCLAIMER_SIZE,
            "disclaimer_y_shift": REELS_SUMMARY_DISCLAIMER_Y_SHIFT,
            "shadow_alpha": REELS_SUMMARY_SHADOW_ALPHA,
            "show_handle": REELS_SUMMARY_SHOW_HANDLE,
            "handle_text": REELS_SUMMARY_HANDLE_TEXT,
        }

    return {
        "figsize": (GENERAL_SUMMARY_FIG_W, GENERAL_SUMMARY_FIG_H),
        "dpi": GENERAL_SUMMARY_DPI,
        "card_x": GENERAL_SUMMARY_CARD_X,
        "card_y": GENERAL_SUMMARY_CARD_Y,
        "card_w": GENERAL_SUMMARY_CARD_W,
        "card_h": GENERAL_SUMMARY_CARD_H,
        "title_y": GENERAL_SUMMARY_TITLE_Y,
        "subtitle_y": GENERAL_SUMMARY_SUBTITLE_Y,
        "footer_y": GENERAL_SUMMARY_FOOTER_Y,
        "title_size": GENERAL_SUMMARY_TITLE_SIZE,
        "subtitle_size": GENERAL_SUMMARY_SUBTITLE_SIZE,
        "strip_text_size": GENERAL_SUMMARY_STRIP_TEXT_SIZE,
        "score_size": GENERAL_SUMMARY_SCORE_SIZE,
        "label_size": GENERAL_SUMMARY_LABEL_SIZE,
        "value_size": GENERAL_SUMMARY_VALUE_SIZE,
        "paragraph_size": GENERAL_SUMMARY_PARAGRAPH_SIZE,
        "line_spacing": GENERAL_SUMMARY_LINE_SPACING,
        "strip_row_gap": GENERAL_SUMMARY_STRIP_ROW_GAP,
        "row_gap": GENERAL_SUMMARY_ROW_GAP,
        "top_strip_label": GENERAL_SUMMARY_TOP_STRIP_LABEL,
        "disclaimer_text": GENERAL_SUMMARY_DISCLAIMER_TEXT,
        "disclaimer_size": GENERAL_SUMMARY_DISCLAIMER_SIZE,
        "disclaimer_y_shift": GENERAL_SUMMARY_DISCLAIMER_Y_SHIFT,
        "shadow_alpha": GENERAL_SUMMARY_SHADOW_ALPHA,
        "show_handle": GENERAL_SUMMARY_SHOW_HANDLE,
        "handle_text": GENERAL_SUMMARY_HANDLE_TEXT,
    }


def _fit_text_to_width(ax, fig, text, max_width_px, fontsize, fontweight="bold", max_lines=2):
    text = "" if text is None else str(text).strip()
    words = text.split()
    if not words:
        return ""

    def _width_px(t):
        temp = ax.text(0, 0, t, fontsize=fontsize, fontweight=fontweight, alpha=0)
        fig.canvas.draw()
        width = temp.get_window_extent(renderer=fig.canvas.get_renderer()).width
        temp.remove()
        return width

    if _width_px(text) <= max_width_px:
        return text

    lines = []
    current = ""
    for word in words:
        candidate = word if not current else current + " " + word
        if _width_px(candidate) <= max_width_px:
            current = candidate
        else:
            if current:
                lines.append(current)
                current = word
            else:
                current = word
            if len(lines) >= max_lines:
                break

    if current and len(lines) < max_lines:
        lines.append(current)

    if lines:
        last = lines[-1]
        while last and _width_px(last + "...") > max_width_px:
            last = last[:-1].rstrip()
        if last != lines[-1]:
            lines[-1] = last + "..."

    return "\n".join(lines[:max_lines])


def _fmt_pct(value, suffix="%") -> str:
    try:
        if pd.isna(value):
            return "N/A"
        return f"{float(value):+.1f}{suffix}"
    except Exception:
        return "N/A"



def format_expected_move(value: float, show_move: str) -> str:
    """
    Same behavior as the original script:
    SHOW_MOVE = NO hides the expected move values as **.
    """
    sm = str(show_move or "YES").strip().upper()
    if sm in ("NO", "N", "FALSE", "0"):
        return "**"
    try:
        return f"{float(value):+.2f}%"
    except Exception:
        return "**"


def fetch_prices(symbol: str, start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    px = yf.download(
        symbol,
        start=start.strftime("%Y-%m-%d"),
        end=end.strftime("%Y-%m-%d"),
        progress=False,
        auto_adjust=True,
    )
    if px is None or px.empty:
        raise RuntimeError(f"No price data fetched for {symbol}. Check symbol.")
    px = px.reset_index()
    px["Date"] = pd.to_datetime(px["Date"]).dt.tz_localize(None)
    return px[["Date", "Close"]].sort_values("Date").reset_index(drop=True)


def next_trading_day(prices: pd.DataFrame, dt: pd.Timestamp) -> pd.Timestamp:
    s = prices.loc[prices["Date"] >= dt, "Date"]
    return pd.NaT if s.empty else s.iloc[0]


def forward_return(prices: pd.DataFrame, anchor: pd.Timestamp, n: int) -> float:
    if pd.isna(anchor):
        return np.nan
    idx = prices.index[prices["Date"] == anchor]
    if len(idx) == 0:
        return np.nan
    i = int(idx[0])
    j = i + n
    if j >= len(prices):
        return np.nan
    c0 = float(prices.loc[i, "Close"])
    c1 = float(prices.loc[j, "Close"])
    return (c1 / c0 - 1.0) * 100.0


def train_predict(df: pd.DataFrame, ycol: str, feat_cols: list[str]) -> tuple[float, dict]:
    """
    Original Ridge-based result-impact model, with the same safety cleaning used earlier.
    """
    d = df.copy()
    for c in feat_cols + [ycol]:
        d[c] = pd.to_numeric(d[c], errors="coerce")
    d = d.replace([np.inf, -np.inf], np.nan)
    d = d.dropna(subset=feat_cols + [ycol]).copy()

    HUGE = 1e12
    if not d.empty:
        arr = d[feat_cols].to_numpy(dtype=np.float64)
        mask_finite = np.isfinite(arr).all(axis=1)
        mask_huge = (np.abs(arr) < HUGE).all(axis=1)
        d = d.loc[mask_finite & mask_huge].copy()

    if len(d) < MIN_TRAIN_EVENTS:
        return np.nan, {
            "status": "insufficient_history",
            "events": int(len(d)),
            "note": "Dropped non-finite/huge rows (inf/-inf/NaN).",
        }

    X = d[feat_cols].to_numpy(dtype=np.float64)
    y = d[ycol].to_numpy(dtype=np.float64)

    if len(d) < 2:
        return np.nan, {"status": "insufficient_history", "events": int(len(d))}

    model = Pipeline([("scaler", StandardScaler()), ("ridge", Ridge(alpha=1.0))])
    model.fit(X[:-1], y[:-1])
    pred = float(model.predict(X[-1:].reshape(1, -1))[0])

    train_pred = model.predict(X[:-1])
    resid = y[:-1] - train_pred
    mae = float(np.mean(np.abs(resid)))
    denom = float(np.sum((y[:-1] - np.mean(y[:-1])) ** 2))
    r2 = float(1 - (np.sum(resid ** 2) / denom)) if denom > 1e-9 else 0.0

    return pred, {"status": "ok", "train_events": int(len(d) - 1), "mae": mae, "r2": r2}


def conf_label(d1: dict, d3: dict) -> str:
    def _score(diag):
        if diag.get("status") != "ok":
            return 0
        n = diag.get("train_events", 0)
        r2 = diag.get("r2", 0)
        score = 0
        if n >= 12:
            score += 2
        elif n >= 8:
            score += 1
        if r2 >= 0.25:
            score += 2
        elif r2 >= 0.10:
            score += 1
        return score

    sc = max(_score(d1), _score(d3))
    if sc >= 4:
        return "High"
    if sc >= 2:
        return "Medium"
    return "Low"


def calculate_result_impact_summary(q: pd.DataFrame, yf_symbol: str) -> dict:
    """
    Restores the original ResultImpactPredictor.py summary calculation:
    - Direction
    - Expected Move 1D %
    - Expected Move 3D %
    - Confidence
    """
    q2 = q.copy()
    yf_symbol = str(yf_symbol).strip().upper()
    if not yf_symbol.endswith((".NS", ".BO")):
        yf_symbol = f"{yf_symbol}.NS"

    q2["EventDate"] = q2["QuarterEnd"] + pd.to_timedelta(FALLBACK_EVENT_OFFSET_DAYS, unit="D")

    start = q2["EventDate"].min() - pd.to_timedelta(PRICE_BUFFER_DAYS, unit="D")
    end = q2["EventDate"].max() + pd.to_timedelta(PRICE_BUFFER_DAYS + 25, unit="D")
    prices = fetch_prices(yf_symbol, start, end)

    q2["AnchorDate"] = q2["EventDate"].apply(lambda d: next_trading_day(prices, d))
    for w in EVENT_WINDOWS:
        q2[f"Ret_{w}D%"] = q2["AnchorDate"].apply(lambda d: forward_return(prices, d, w))

    f = build_features(q2)
    feat_cols = [
        "SurpriseScore",
        "NP_YoY%", "EPS_YoY%", "Rev_YoY%",
        "NP_QoQ%", "EPS_QoQ%", "Rev_QoQ%",
        "NP_Margin%", "NP_Margin_Delta_YoY", "NP_Margin_Delta_QoQ",
        "Exp_YoY%", "Exp_QoQ%",
    ]

    pred1, d1 = train_predict(f, "Ret_1D%", feat_cols)
    pred3, d3 = train_predict(f, "Ret_3D%", feat_cols)
    latest = f.iloc[-1]

    if np.isnan(pred1):
        pred1 = float(np.clip(latest["SurpriseScore"] * 0.9, -3.0, 3.0))
    if np.isnan(pred3):
        pred3 = float(np.clip(latest["SurpriseScore"] * 1.6, -6.0, 6.0))

    direction = "UP" if (pred1 + pred3) >= 0 else "DOWN"
    confidence = conf_label(d1, d3)

    drivers = []
    def _add_driver(name, val):
        if pd.notna(val) and abs(float(val)) >= 8:
            drivers.append(f"{name} {float(val):+.1f}%")

    _add_driver("NetProfit YoY", latest.get("NP_YoY%"))
    _add_driver("EPS YoY", latest.get("EPS_YoY%"))
    _add_driver("Revenue YoY", latest.get("Rev_YoY%"))
    _add_driver("Expenses YoY", latest.get("Exp_YoY%"))
    if abs(float(latest.get("NP_Margin_Delta_YoY", 0) or 0)) >= 0.5:
        drivers.append(f"Margin ΔYoY {float(latest['NP_Margin_Delta_YoY']):+.2f} pp")

    driver_txt = ", ".join(drivers) if drivers else "Mixed YoY/QoQ signals"
    impact = "Bullish" if direction == "UP" else "Bearish"
    summary = (
        f"{impact} result-impact estimate for {yf_symbol}. "
        f"Drivers: {driver_txt}. "
        f"Expected move: ~{pred1:+.2f}% (1D), ~{pred3:+.2f}% (3D). "
        f"Confidence: {confidence}."
    )

    return {
        "Direction": direction,
        "ExpectedMove1D%": format_expected_move(pred1, SHOW_MOVE),
        "ExpectedMove3D%": format_expected_move(pred3, SHOW_MOVE),
        "Confidence": confidence,
        "ImpactDriversText": driver_txt,
        "ImpactSummary": summary,
        "Diagnostics": {"1D": d1, "3D": d3},
    }

def build_result_summary_payload(
    fdf: pd.DataFrame,
    score_df: pd.DataFrame,
    latest_quarter: str,
    impact_summary: dict | None = None,
) -> dict:
    latest = fdf.iloc[-1].copy()
    impact_summary = impact_summary or {}
    total_row = score_df.loc[score_df["Parameter"].astype(str).str.upper() == "TOTAL"]

    total_pct = ""
    grade = ""
    if not total_row.empty:
        total_pct = total_row.iloc[0].get("Total%", "")
        grade = str(total_row.iloc[0].get("Grade", "")).strip()

    try:
        score_text = f"{float(total_pct):.1f}%"
    except Exception:
        score_text = "N/A"

    metric_items = [
        ("Revenue YoY", latest.get("Rev_YoY%", np.nan)),
        ("Net Profit YoY", latest.get("NP_YoY%", np.nan)),
        ("EPS YoY", latest.get("EPS_YoY%", np.nan)),
        ("NP Margin", latest.get("NP_Margin%", np.nan)),
        ("Margin ΔYoY", latest.get("NP_Margin_Delta_YoY", np.nan)),
    ]

    metric_rows = []
    for label, value in metric_items:
        if pd.notna(value):
            suffix = " pp" if "Δ" in label else "%"
            metric_rows.append(f"{label}: {_fmt_pct(value, suffix)}")

    impact_rows = [
        f"Direction: {impact_summary.get('Direction', 'N/A')}",
        f"Expected Move 1D %: {impact_summary.get('ExpectedMove1D%', 'N/A')}",
        f"Expected Move 3D %: {impact_summary.get('ExpectedMove3D%', 'N/A')}",
        f"Confidence: {impact_summary.get('Confidence', 'N/A')}",
    ]

    top_rows = score_df[score_df["Parameter"].astype(str).str.upper() != "TOTAL"].copy()
    top_rows["Marks"] = pd.to_numeric(top_rows["Marks"], errors="coerce")
    top_rows = top_rows.sort_values("Marks", ascending=False).head(3)
    strengths = ", ".join(top_rows["Parameter"].astype(str).tolist()) if not top_rows.empty else "Mixed result quality"

    verdict, _ = _summary_grade_to_verdict_and_color(grade)
    summary_text = impact_summary.get("ImpactSummary") or (
        f"{latest_quarter} score stands at {score_text}. "
        f"Overall result tone is {verdict.title()}. "
        f"Key support comes from {strengths}."
    )

    return {
        "LatestQuarter": latest_quarter,
        "ScoreText": score_text,
        "Grade": grade,
        "Verdict": verdict,
        "ExpectedPattern": derive_expected_pattern({"ImpactRows": impact_rows}),
        "ImpactRows": impact_rows,
        "Drivers": metric_rows[:5],
        "Strengths": strengths,
        "Summary": summary_text,
    }


def _parse_summary_row_value(rows, row_label: str) -> str:
    """Return the value part from rows like 'Direction: DOWN'."""
    row_label = str(row_label).strip().lower()
    for item in rows or []:
        if ":" not in str(item):
            continue
        label, value = str(item).split(":", 1)
        if label.strip().lower() == row_label:
            return value.strip()
    return ""


def _to_float_or_nan(value) -> float:
    try:
        s = str(value).strip().replace("%", "").replace("**", "")
        if not s:
            return np.nan
        return float(s)
    except Exception:
        return np.nan


def derive_expected_pattern(summary_payload: dict) -> str:
    """
    Converts the restored impact rows into a clean top-badge pattern label.

    Output examples:
    - Downside Bias
    - Upside Bias
    - Mixed Pattern
    - Volatile
    - Range-bound
    """
    impact_rows = summary_payload.get("ImpactRows", []) or []
    direction = _parse_summary_row_value(impact_rows, "Direction").upper()
    move_1d = _to_float_or_nan(_parse_summary_row_value(impact_rows, "Expected Move 1D %"))
    move_3d = _to_float_or_nan(_parse_summary_row_value(impact_rows, "Expected Move 3D %"))

    moves = [v for v in [move_1d, move_3d] if pd.notna(v)]

    if moves:
        max_abs = max(abs(v) for v in moves)
        avg_move = float(np.mean(moves))

        # Very small estimated movement is better shown as range-bound,
        # even if the numeric direction is slightly UP/DOWN.
        if max_abs <= 0.30:
            return "Range-bound"

        # Opposite signs indicate no clean direction.
        if len(moves) >= 2 and (move_1d * move_3d) < 0:
            return "Mixed Pattern"

        # Large expected moves should be shown as volatile.
        if max_abs >= 5.0:
            return "Volatile"

        if avg_move > 0:
            return "Upside Bias"
        if avg_move < 0:
            return "Downside Bias"

    # Fallback when SHOW_MOVE = NO or values are unavailable.
    if direction == "UP":
        return "Upside Bias"
    if direction == "DOWN":
        return "Downside Bias"
    return "Mixed Pattern"


def _score_text_to_float(score_text) -> float:
    """Convert values like '74.2%' into float for pattern color intensity."""
    try:
        s = str(score_text).strip().replace("%", "")
        return float(s)
    except Exception:
        return np.nan


def _pattern_intensity_from_score(score_text) -> str:
    """
    Score-based strength bucket used only for color depth.
    Higher score = stronger color intensity; lower score = lighter color.
    """
    score = _score_text_to_float(score_text)
    if pd.isna(score):
        return "medium"
    if score >= 75:
        return "strong"
    if score >= 55:
        return "medium"
    return "light"


def get_expected_pattern_colors(expected_pattern: str, score_text: str = "") -> dict:
    """
    Returns pattern-aware colors for the summary strip and badge.

    Rules:
    - Downside Bias => red family; score controls light/medium/strong red.
    - Upside Bias => green family; score controls light/medium/strong green.
    - Mixed => amber/orange family.
    - Volatile => purple/orange warning family.
    - Range-bound => slate/blue-gray family.
    """
    pattern = str(expected_pattern or "").strip().lower()
    intensity = _pattern_intensity_from_score(score_text)

    if "downside" in pattern:
        palette = {
            "light":  {"strip": "#EF4444", "badge": "#FCA5A5", "value": "#991B1B"},
            "medium": {"strip": "#DC2626", "badge": "#F87171", "value": "#7F1D1D"},
            "strong": {"strip": "#B91C1C", "badge": "#EF4444", "value": "#7F1D1D"},
        }
        return palette[intensity]

    if "upside" in pattern:
        palette = {
            "light":  {"strip": "#22C55E", "badge": "#86EFAC", "value": "#166534"},
            "medium": {"strip": "#16A34A", "badge": "#4ADE80", "value": "#14532D"},
            "strong": {"strip": "#15803D", "badge": "#22C55E", "value": "#14532D"},
        }
        return palette[intensity]

    if "volatile" in pattern:
        return {"strip": "#7C3AED", "badge": "#A78BFA", "value": "#4C1D95"}

    if "range" in pattern:
        return {"strip": "#64748B", "badge": "#94A3B8", "value": "#334155"}

    # Mixed Pattern or unknown
    return {"strip": "#D97706", "badge": "#FBBF24", "value": "#92400E"}


def export_result_summary_as_image(
    symbol: str,
    company_name: str,
    summary_payload: dict,
    out_path_template: str,
    image_type: str = "general",
) -> str:
    import os
    import textwrap
    import matplotlib.pyplot as plt
    from matplotlib.patches import FancyBboxPatch

    out_path = out_path_template.replace("{SYMBOL}", symbol)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    cfg = _get_summary_image_layout(image_type)

    BG_NAVY = "#0B1E3A"
    GOLD_TITLE = "#E5C26A"
    TEXT_LIGHT = "#E8EDF4"
    CARD_BG = "#E9EDF2"
    CARD_EDGE = "#D6D9DE"
    LABEL_DARK = "#334155"
    VALUE_DARK = "#0F172A"
    MUTED_DARK = "#475569"
    PANEL_BG = "#F8FAFC"
    GRID_BORDER = "#CBD5E1"

    verdict = summary_payload.get("Verdict", "SOFT")
    expected_pattern = summary_payload.get("ExpectedPattern") or derive_expected_pattern(summary_payload)

    # Pattern-aware colors: Downside = red family, Upside = green family,
    # Mixed/Volatile/Range-bound = neutral/warning families.
    # Color intensity is controlled by ScoreText, but score is not printed here.
    pattern_colors = get_expected_pattern_colors(
        expected_pattern=expected_pattern,
        score_text=summary_payload.get("ScoreText", ""),
    )
    strip_color = pattern_colors["strip"]
    badge_color = pattern_colors["badge"]

    latest_quarter = summary_payload.get("LatestQuarter", "")
    title_suffix = f"{latest_quarter} Result Summary" if latest_quarter else "Result Summary"

    fig = plt.figure(figsize=cfg["figsize"], dpi=cfg["dpi"])
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_axis_off()
    fig.patch.set_facecolor(BG_NAVY)

    card_x, card_y, card_w, card_h = cfg["card_x"], cfg["card_y"], cfg["card_w"], cfg["card_h"]
    max_title_width_px = fig.bbox.width * card_w * 0.98
    fitted_company_name = _fit_text_to_width(
        ax, fig, company_name,
        max_width_px=max_title_width_px,
        fontsize=cfg["title_size"],
        fontweight="bold",
        max_lines=2,
    )
    title_lines = fitted_company_name.count("\n") + 1
    title_y = cfg["title_y"] + (0.018 if title_lines > 1 else 0)
    subtitle_y = cfg["subtitle_y"] - (0.010 if title_lines > 1 else 0)

    ax.text(
        0.5, title_y, fitted_company_name,
        ha="center", va="center",
        fontsize=cfg["title_size"],
        fontweight="bold",
        color=GOLD_TITLE,
        linespacing=0.92 if title_lines > 1 else 1.05,
    )
    ax.text(
        0.5, subtitle_y, title_suffix,
        ha="center", va="center",
        fontsize=cfg["subtitle_size"],
        color=TEXT_LIGHT,
        alpha=0.90,
    )

    shadow = FancyBboxPatch(
        (card_x, card_y - 0.006), card_w, card_h,
        boxstyle="round,pad=0.012,rounding_size=0.03",
        linewidth=0, facecolor="black", alpha=cfg["shadow_alpha"],
        transform=ax.transAxes, zorder=0,
    )
    ax.add_patch(shadow)

    card = FancyBboxPatch(
        (card_x, card_y), card_w, card_h,
        boxstyle="round,pad=0.012,rounding_size=0.03",
        linewidth=1.5, edgecolor=CARD_EDGE, facecolor=CARD_BG,
        transform=ax.transAxes, zorder=1,
    )
    ax.add_patch(card)

    strip_h = 0.075 if image_type != "reels" else 0.065
    strip_y = card_y + card_h - strip_h - 0.022
    strip = FancyBboxPatch(
        (card_x + 0.025, strip_y), card_w - 0.050, strip_h,
        boxstyle="round,pad=0.010,rounding_size=0.02",
        linewidth=0, facecolor=strip_color,
        transform=ax.transAxes, zorder=3,
    )
    ax.add_patch(strip)

    top_strip_label = str(cfg.get("top_strip_label", "IMPACT ESTIMATE")).strip() or "IMPACT ESTIMATE"
    ax.text(
        card_x + 0.052, strip_y + strip_h / 2,
        top_strip_label.upper(),
        transform=ax.transAxes, ha="left", va="center",
        fontsize=cfg["strip_text_size"], fontweight="bold",
        color="white", zorder=4,
    )

    # Top-right badge now shows the expected pattern, not the result-quality score.
    # This avoids implying that the score is the probability of the impact estimate.
    badge_w = 0.31 if image_type != "reels" else 0.34
    badge_h = strip_h * 0.76
    badge_x = card_x + card_w - badge_w - 0.040
    badge_y = strip_y + (strip_h - badge_h) / 2
    badge = FancyBboxPatch(
        (badge_x, badge_y), badge_w, badge_h,
        boxstyle="round,pad=0.008,rounding_size=0.03",
        linewidth=0, facecolor=badge_color, alpha=0.38,
        transform=ax.transAxes, zorder=5,
    )
    ax.add_patch(badge)

    badge_font_size = max(cfg["label_size"], cfg["strip_text_size"] - 5)
    if len(str(expected_pattern)) >= 13:
        badge_font_size = max(cfg["label_size"] - 1, cfg["strip_text_size"] - 7)

    ax.text(
        badge_x + badge_w / 2, badge_y + badge_h / 2,
        str(expected_pattern),
        transform=ax.transAxes, ha="center", va="center",
        fontsize=badge_font_size,
        fontweight="bold", color="white", zorder=6,
    )

    # Summary rows
    # Score panel remains removed, but the original calculated summary fields are restored.
    impact_rows = summary_payload.get("ImpactRows", []) or []
    metric_rows = summary_payload.get("Drivers", []) or []
    display_rows = impact_rows + metric_rows[:5]

    row_count = max(1, len(display_rows))
    if row_count >= 9:
        row_h = 0.038 if image_type != "reels" else 0.034
    else:
        row_h = 0.052 if image_type != "reels" else 0.047

    # Button-controlled gap/padding between rows 1, 2, 3 ... and the summary paragraph box.
    # Increase *_SUMMARY_ROW_GAP to create more vertical space between rows.
    # Decrease it if rows need to fit tighter inside the card.
    row_gap = float(cfg.get("row_gap", 0.010 if image_type != "reels" else 0.008))

    # Button-controlled padding between the colored result strip and first row.
    # Increase *_SUMMARY_STRIP_ROW_GAP to move rows downward.
    # Decrease it to move rows upward.
    row_start_y = strip_y - float(cfg.get("strip_row_gap", 0.030 if image_type != "reels" else 0.028))

    last_row_bottom = row_start_y
    for idx, item in enumerate(display_rows):
        y = row_start_y - idx * (row_h + row_gap)
        row_bottom = y - row_h
        last_row_bottom = row_bottom

        row_box = FancyBboxPatch(
            (card_x + 0.045, row_bottom), card_w - 0.090, row_h,
            boxstyle="round,pad=0.006,rounding_size=0.014",
            linewidth=0.8, edgecolor=GRID_BORDER, facecolor=PANEL_BG,
            transform=ax.transAxes, zorder=3,
        )
        ax.add_patch(row_box)

        if ":" in item:
            label, value = item.split(":", 1)
        else:
            label, value = item, ""

        ax.text(
            card_x + 0.065, row_bottom + row_h / 2,
            label.strip(),
            transform=ax.transAxes, ha="left", va="center",
            fontsize=cfg["label_size"], fontweight="bold",
            color=LABEL_DARK, zorder=4,
        )
        value_clean = value.strip()
        value_color = VALUE_DARK
        label_clean = label.strip().lower()
        if label_clean == "direction":
            if value_clean.upper() == "UP":
                value_color = "#15803D"
            elif value_clean.upper() == "DOWN":
                value_color = "#991B1B"
        elif label_clean == "confidence":
            if value_clean.title() == "High":
                value_color = "#15803D"
            elif value_clean.title() == "Medium":
                value_color = "#9C6500"
            elif value_clean:
                value_color = "#991B1B"

        ax.text(
            card_x + card_w - 0.065, row_bottom + row_h / 2,
            value_clean,
            transform=ax.transAxes, ha="right", va="center",
            fontsize=cfg["value_size"], fontweight="bold",
            color=value_color, zorder=4,
        )

    # The old lower paragraph/conclusion box has been removed intentionally.
    # Its content duplicated Direction / Expected Move / Confidence / driver rows above
    # and created crowding in summary images.

    # Brief disclaimer inside the remaining lower card space.
    # Use *_SUMMARY_DISCLAIMER_TEXT / *_SUMMARY_DISCLAIMER_SIZE /
    # *_SUMMARY_DISCLAIMER_Y_SHIFT buttons to control it per image type.
    disclaimer_text = str(cfg.get("disclaimer_text", "")).strip()
    if disclaimer_text:
        disclaimer_y = card_y + 0.032 + float(cfg.get("disclaimer_y_shift", 0.0))
        ax.text(
            card_x + card_w / 2, disclaimer_y, disclaimer_text,
            transform=ax.transAxes, ha="center", va="center",
            fontsize=cfg.get("disclaimer_size", 8.0),
            fontweight="bold", color=MUTED_DARK, alpha=0.90, zorder=4,
        )

    if cfg.get("show_handle", False):
        fig.text(
            0.5, cfg["footer_y"], cfg.get("handle_text", "automationintrade"),
            ha="center", va="center",
            fontsize=12, color=TEXT_LIGHT, fontweight="bold",
        )

    plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
    fig.savefig(out_path, bbox_inches="tight", pad_inches=0)
    plt.close(fig)
    return out_path



def _safe_filename_part(value: str) -> str:
    value = str(value or "").strip()
    value = re.sub(r"[^A-Za-z0-9_\-]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value or "Frame"


def _reel_metric_value(summary_payload: dict, label: str) -> str:
    return _parse_summary_row_value(summary_payload.get("ImpactRows", []) or [], label)


def _shorten_metric_label(label: str) -> str:
    mapping = {
        "expected move 1d %": "1D Move",
        "expected move 3d %": "3D Move",
        "netprofit yoy%": "Profit YoY",
        "netprofit qoq%": "Profit QoQ",
        "revenue yoy%": "Revenue YoY",
        "revenue qoq%": "Revenue QoQ",
        "eps yoy%": "EPS YoY",
        "eps qoq%": "EPS QoQ",
        "margin Δyoy (pp)": "Margin YoY",
        "margin Δqoq (pp)": "Margin QoQ",
    }
    key = str(label or "").strip().lower()
    return mapping.get(key, str(label or "").strip())


def _split_label_value(row: str) -> tuple[str, str]:
    if ":" in str(row):
        label, value = str(row).split(":", 1)
        return label.strip(), value.strip()
    return str(row).strip(), ""


def _draw_reel_card(ax, x, y, w, h, facecolor="#F8FAFC", edgecolor="#CBD5E1", zorder=2):
    from matplotlib.patches import FancyBboxPatch
    card = FancyBboxPatch(
        (x, y), w, h,
        boxstyle="round,pad=0.014,rounding_size=0.030",
        linewidth=1.0, edgecolor=edgecolor, facecolor=facecolor,
        transform=ax.transAxes, zorder=zorder,
    )
    ax.add_patch(card)
    return card


def _reel_text_width_px(ax, text, fontsize, fontweight="bold"):
    """Return rendered text width in pixels for safe fitting."""
    fig = ax.figure
    temp = ax.text(0, 0, str(text), fontsize=fontsize, fontweight=fontweight, alpha=0)
    fig.canvas.draw()
    width = temp.get_window_extent(renderer=fig.canvas.get_renderer()).width
    temp.remove()
    return width


def _wrap_text_to_px(ax, text, max_width_px, fontsize, fontweight="bold", max_lines=2):
    """Wrap text by words so it stays inside the available pixel width."""
    import textwrap

    text = "" if text is None else str(text).strip()
    if not text:
        return ""

    if _reel_text_width_px(ax, text, fontsize, fontweight) <= max_width_px:
        return text

    words = text.split()
    lines = []
    current = ""
    for word in words:
        candidate = word if not current else f"{current} {word}"
        if _reel_text_width_px(ax, candidate, fontsize, fontweight) <= max_width_px:
            current = candidate
        else:
            if current:
                lines.append(current)
                current = word
            else:
                # Very long single word: character-wrap it instead of allowing overflow.
                chunks = textwrap.wrap(word, width=max(6, int(max_width_px / max(fontsize * 0.62, 1))))
                lines.extend(chunks[:max_lines])
                current = ""
            if len(lines) >= max_lines:
                break
    if current and len(lines) < max_lines:
        lines.append(current)

    if not lines:
        lines = [text]

    # Trim final line safely if even wrapped text is too wide.
    final = lines[:max_lines]
    last = final[-1]
    while last and _reel_text_width_px(ax, last + "...", fontsize, fontweight) > max_width_px:
        last = last[:-1].rstrip()
    if last != final[-1]:
        final[-1] = (last + "...") if last else "..."
    return "\n".join(final)


def _fit_fontsize_to_px(ax, text, max_width_px, start_size, min_size=9, fontweight="bold", max_lines=2):
    """Reduce font size and wrap text until it fits inside max_width_px."""
    size = float(start_size)
    while size >= float(min_size):
        wrapped = _wrap_text_to_px(ax, text, max_width_px, size, fontweight=fontweight, max_lines=max_lines)
        widest = max(wrapped.split("\n"), key=len) if wrapped else ""
        if _reel_text_width_px(ax, widest, size, fontweight) <= max_width_px:
            return wrapped, size
        size -= 1
    wrapped = _wrap_text_to_px(ax, text, max_width_px, min_size, fontweight=fontweight, max_lines=max_lines)
    return wrapped, float(min_size)


def _draw_reel_center_text(ax, x, y, text, max_w_frac, fontsize, color, fontweight="bold", max_lines=2, zorder=4, min_size=9, va="center"):
    """Draw centered text that auto-shrinks/wraps inside max_w_frac of figure width."""
    max_width_px = ax.figure.bbox.width * max_w_frac
    fitted, fitted_size = _fit_fontsize_to_px(
        ax, text, max_width_px=max_width_px, start_size=fontsize,
        min_size=min_size, fontweight=fontweight, max_lines=max_lines,
    )
    ax.text(
        x, y, fitted,
        transform=ax.transAxes, ha="center", va=va,
        fontsize=fitted_size, fontweight=fontweight, color=color,
        linespacing=0.95, zorder=zorder,
    )


def _draw_reel_row(
    ax, x, y, w, h, label, value,
    label_color="#334155", value_color="#0F172A",
    label_size=15, value_size=17,
    value_area_ratio=0.56,
    max_value_lines=2,
):
    """
    Draw one rounded metric row.
    Fixes overflow by giving label/value separate safe areas and auto-fitting both texts.
    """
    _draw_reel_card(ax, x, y, w, h, facecolor="#F8FAFC", edgecolor="#CBD5E1", zorder=3)

    fig_w_px = ax.figure.bbox.width
    left_pad = 0.030
    right_pad = 0.030
    gap = 0.018
    value_area_w = w * value_area_ratio
    label_area_w = max(0.16, w - value_area_w - left_pad - right_pad - gap)

    label_text, fitted_label_size = _fit_fontsize_to_px(
        ax, str(label), fig_w_px * label_area_w,
        start_size=label_size, min_size=9,
        fontweight="bold", max_lines=1,
    )
    value_text, fitted_value_size = _fit_fontsize_to_px(
        ax, str(value), fig_w_px * value_area_w,
        start_size=value_size, min_size=9,
        fontweight="bold", max_lines=max_value_lines,
    )

    ax.text(
        x + left_pad, y + h / 2, label_text,
        transform=ax.transAxes, ha="left", va="center",
        fontsize=fitted_label_size, fontweight="bold", color=label_color,
        linespacing=0.92, zorder=4,
    )
    ax.text(
        x + w - right_pad, y + h / 2, value_text,
        transform=ax.transAxes, ha="right", va="center",
        fontsize=fitted_value_size, fontweight="bold", color=value_color,
        linespacing=0.92, zorder=4,
    )


def _reel_value_color(label: str, value: str) -> str:
    label_clean = str(label or "").lower().strip()
    value_clean = str(value or "").upper().strip()
    if label_clean == "direction":
        if value_clean == "UP":
            return "#15803D"
        if value_clean == "DOWN":
            return "#991B1B"
    if label_clean == "confidence":
        if value_clean.title() == "High":
            return "#15803D"
        if value_clean.title() == "Medium":
            return "#9C6500"
        return "#991B1B"
    try:
        num = float(str(value).replace("%", "").replace("+", "").strip())
        if num > 0:
            return "#15803D"
        if num < 0:
            return "#991B1B"
    except Exception:
        pass
    return "#0F172A"


def export_result_reel_frame_as_image(
    symbol: str,
    company_name: str,
    summary_payload: dict,
    frame_no: int,
    frame_name: str,
    frame_type: str,
    out_path_template: str = REELS_FRAME_PACK_OUTPUT_TEMPLATE,
) -> str:
    """
    Generate one Instagram Reel-friendly 9:16 frame.

    Frame types:
    - hook    : first frame; instantly explains why the Reel is worth watching.
    - impact  : compact direction / expected move / confidence snapshot.
    - drivers : top result drivers, lighter than the heavy table image.
    - setup   : watchlist/trading setup style closing frame.
    """
    import os
    import textwrap
    import matplotlib.pyplot as plt
    from matplotlib.patches import FancyBboxPatch

    safe_frame_name = _safe_filename_part(frame_name)
    out_path = (
        out_path_template
        .replace("{SYMBOL}", str(symbol))
        .replace("{FRAME_NO}", f"{int(frame_no):02d}")
        .replace("{FRAME_NAME}", safe_frame_name)
    )
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    BG_NAVY = "#0B1E3A"
    GOLD_TITLE = "#E5C26A"
    TEXT_LIGHT = "#E8EDF4"
    MUTED_LIGHT = "#B6C2D2"
    CARD_BG = "#E9EDF2"
    LABEL_DARK = "#334155"
    VALUE_DARK = "#0F172A"
    MUTED_DARK = "#475569"

    expected_pattern = summary_payload.get("ExpectedPattern") or derive_expected_pattern(summary_payload)
    pattern_colors = get_expected_pattern_colors(
        expected_pattern=expected_pattern,
        score_text=summary_payload.get("ScoreText", ""),
    )
    strip_color = pattern_colors["strip"]
    badge_color = pattern_colors["badge"]
    latest_quarter = summary_payload.get("LatestQuarter", "")
    score_text = str(summary_payload.get("ScoreText", "")).strip()
    direction = _reel_metric_value(summary_payload, "Direction") or "N/A"
    move_1d = _reel_metric_value(summary_payload, "Expected Move 1D %") or "N/A"
    move_3d = _reel_metric_value(summary_payload, "Expected Move 3D %") or "N/A"
    confidence = _reel_metric_value(summary_payload, "Confidence") or "N/A"

    fig = plt.figure(figsize=(REELS_FRAME_FIG_W, REELS_FRAME_FIG_H), dpi=REELS_FRAME_DPI)
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_axis_off()
    fig.patch.set_facecolor(BG_NAVY)

    # Tiny top brand/site cue.
    ax.text(
        0.5, 0.965, "automationintrade.com",
        transform=ax.transAxes, ha="center", va="center",
        fontsize=13, fontweight="bold", color=MUTED_LIGHT, alpha=0.95,
    )

    # Company title, auto-wrapped to 2 lines.
    fitted_company = _fit_text_to_width(
        ax, fig, company_name,
        max_width_px=fig.bbox.width * 0.86,
        fontsize=23,
        fontweight="bold",
        max_lines=2,
    )
    ax.text(
        0.5, 0.910, fitted_company,
        transform=ax.transAxes, ha="center", va="center",
        fontsize=23, fontweight="bold", color=GOLD_TITLE,
        linespacing=0.94,
    )
    ax.text(
        0.5, 0.862, f"{latest_quarter} Result" if latest_quarter else "Result Update",
        transform=ax.transAxes, ha="center", va="center",
        fontsize=14, fontweight="bold", color=TEXT_LIGHT, alpha=0.92,
    )

    # Main white card.
    shadow = FancyBboxPatch(
        (0.075, 0.145), 0.85, 0.665,
        boxstyle="round,pad=0.018,rounding_size=0.040",
        linewidth=0, facecolor="black", alpha=0.13,
        transform=ax.transAxes, zorder=0,
    )
    ax.add_patch(shadow)
    _draw_reel_card(ax, 0.070, 0.153, 0.86, 0.665, facecolor=CARD_BG, edgecolor="#D6D9DE", zorder=1)

    if frame_type == "hook":
        _draw_reel_center_text(
            ax, 0.5, 0.745, REELS_FRAME_01_HOOK_TITLE,
            max_w_frac=0.78, fontsize=23, color=VALUE_DARK,
            max_lines=2, min_size=14, zorder=3,
        )
        _draw_reel_center_text(
            ax, 0.5, 0.690, REELS_FRAME_01_HOOK_SUBTITLE,
            max_w_frac=0.74, fontsize=13.5, color=MUTED_DARK,
            max_lines=2, min_size=9, zorder=3,
        )
        # Big pattern pill: viewer understands the point immediately.
        pill = FancyBboxPatch(
            (0.145, 0.520), 0.71, 0.120,
            boxstyle="round,pad=0.018,rounding_size=0.045",
            linewidth=0, facecolor=strip_color,
            transform=ax.transAxes, zorder=3,
        )
        ax.add_patch(pill)
        _draw_reel_center_text(
            ax, 0.5, 0.580, str(expected_pattern).upper(),
            max_w_frac=0.62, fontsize=24, color="white",
            max_lines=2, min_size=14, zorder=4,
        )
        _draw_reel_center_text(
            ax, 0.5, 0.448, "Watch this before making a quick trade decision",
            max_w_frac=0.76, fontsize=15, color=VALUE_DARK,
            max_lines=2, min_size=9, zorder=3,
        )
        _draw_reel_row(ax, 0.155, 0.335, 0.69, 0.062, "Direction", direction, value_color=_reel_value_color("Direction", direction))
        _draw_reel_row(ax, 0.155, 0.255, 0.69, 0.062, "Confidence", confidence, value_color=_reel_value_color("Confidence", confidence))

    elif frame_type == "impact":
        _draw_reel_center_text(
            ax, 0.5, 0.742, "IMPACT SNAPSHOT",
            max_w_frac=0.76, fontsize=24, color=VALUE_DARK,
            max_lines=1, min_size=13, zorder=3,
        )
        _draw_reel_center_text(
            ax, 0.5, 0.700, str(expected_pattern),
            max_w_frac=0.76, fontsize=16, color=strip_color,
            max_lines=1, min_size=10, zorder=3,
        )
        rows = [
            ("Direction", direction),
            ("1D Move", move_1d),
            ("3D Move", move_3d),
            ("Confidence", confidence),
        ]
        y = 0.585
        for label, value in rows:
            _draw_reel_row(ax, 0.135, y, 0.73, 0.068, label, value, value_color=_reel_value_color(label, value), label_size=15, value_size=18)
            y -= 0.092
        _draw_reel_center_text(
            ax, 0.5, 0.215, "Use this as a reaction-risk filter, not a blind buy/sell call.",
            max_w_frac=0.78, fontsize=13, color=MUTED_DARK,
            max_lines=2, min_size=9, zorder=3,
        )

    elif frame_type == "drivers":
        _draw_reel_center_text(
            ax, 0.5, 0.742, "WHY THIS PATTERN?",
            max_w_frac=0.76, fontsize=24, color=VALUE_DARK,
            max_lines=1, min_size=13, zorder=3,
        )
        _draw_reel_center_text(
            ax, 0.5, 0.700, "Top result drivers simplified",
            max_w_frac=0.76, fontsize=14.5, color=MUTED_DARK,
            max_lines=1, min_size=10, zorder=3,
        )
        driver_rows = summary_payload.get("Drivers", []) or []
        if not driver_rows:
            driver_rows = ["Result Quality: Mixed", "Trend: Needs confirmation", "Reaction Risk: Watch first candle"]
        y = 0.585
        for raw in driver_rows[:4]:
            label, value = _split_label_value(raw)
            label = _shorten_metric_label(label)
            _draw_reel_row(ax, 0.115, y, 0.77, 0.070, label, value or "N/A", value_color=_reel_value_color(label, value), label_size=14, value_size=16)
            y -= 0.094
        _draw_reel_center_text(
            ax, 0.5, 0.205, "This frame gives the reason, not just the headline.",
            max_w_frac=0.78, fontsize=13, color=MUTED_DARK,
            max_lines=2, min_size=9, zorder=3,
        )

    else:  # setup
        _draw_reel_center_text(
            ax, 0.5, 0.742, "TRADING SETUP VIEW",
            max_w_frac=0.78, fontsize=24, color=VALUE_DARK,
            max_lines=1, min_size=13, zorder=3,
        )
        setup_lines = [
            ("Mood", expected_pattern),
            ("Reaction", "Wait for price confirmation"),
            ("Risk", "Avoid chasing first spike"),
            ("Score", score_text or "N/A"),
        ]
        y = 0.615
        for label, value in setup_lines:
            _draw_reel_row(
                ax, 0.105, y, 0.79, 0.080, label, value,
                value_color=(strip_color if label in ("Mood", "Score") else VALUE_DARK),
                label_size=14, value_size=15, value_area_ratio=0.64, max_value_lines=2,
            )
            y -= 0.103
        _draw_reel_center_text(
            ax, 0.5, 0.205, "Save this result map and compare it with actual price reaction.",
            max_w_frac=0.78, fontsize=12.8, color=MUTED_DARK,
            max_lines=2, min_size=8.5, zorder=3,
        )

    # Bottom disclaimer and handle.
    if str(REELS_FRAME_DISCLAIMER_TEXT).strip():
        ax.text(
            0.5, REELS_FRAME_DISCLAIMER_Y, REELS_FRAME_DISCLAIMER_TEXT,
            transform=ax.transAxes, ha="center", va="center",
            fontsize=9.5, fontweight="bold", color=TEXT_LIGHT, alpha=0.80,
        )
    if REELS_FRAME_SHOW_HANDLE:
        ax.text(
            0.5, REELS_FRAME_FOOTER_Y, REELS_FRAME_HANDLE_TEXT,
            transform=ax.transAxes, ha="center", va="center",
            fontsize=13, fontweight="bold", color=TEXT_LIGHT,
        )

    plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
    fig.savefig(out_path, bbox_inches="tight", pad_inches=0)
    plt.close(fig)
    return out_path


def generate_reels_frame_pack(symbol: str, company_name: str, summary_payload: dict) -> dict:
    """
    Creates multiple lightweight Reel frames:
    1. Hook frame: viewer immediately understands why to watch.
    2. Impact frame: direction/move/confidence.
    3. Drivers frame: simplified reason behind the result mood.
    4. Setup frame: practical next-watch logic.
    """
    frame_specs = [
        (1, "Hook", "hook"),
        (2, "Impact", "impact"),
        (3, "Drivers", "drivers"),
        (4, "Setup", "setup"),
    ]

    paths = {}
    for frame_no, frame_name, frame_type in frame_specs:
        key = f"ReelsFrame{frame_no:02d}{frame_name}"
        path = export_result_reel_frame_as_image(
            symbol=symbol,
            company_name=company_name,
            summary_payload=summary_payload,
            frame_no=frame_no,
            frame_name=frame_name,
            frame_type=frame_type,
            out_path_template=REELS_FRAME_PACK_OUTPUT_TEMPLATE,
        )
        paths[key] = path
        duration_key = f"Frame_{frame_no:02d}_{frame_name}"
        duration = REELS_FRAME_DURATION_HINT_SECONDS.get(duration_key)
        if duration:
            print(f"Saved Reel Frame {frame_no}: {path} | Suggested duration: {duration}s")
        else:
            print(f"Saved Reel Frame {frame_no}: {path}")

    return paths


def _extract_total_score_text(score_df: pd.DataFrame) -> str:
    """
    Extracts latest score from the TOTAL row so the VideoFrames score
    exactly matches the score shown on the General scorecard badge.
    """
    try:
        trow = score_df.loc[score_df["Parameter"].astype(str).str.upper() == "TOTAL"]
        if not trow.empty:
            total_pct = trow.iloc[0].get("Total%", "")
            if total_pct != "" and pd.notna(total_pct):
                return f"{float(total_pct):.1f}%"
    except Exception:
        pass
    return ""


def _extract_total_grade_text(score_df: pd.DataFrame) -> str:
    """
    Extracts the grade from the TOTAL row so VideoFrames can use the same
    condition-based color logic as the top result strip.
    """
    try:
        trow = score_df.loc[score_df["Parameter"].astype(str).str.upper() == "TOTAL"]
        if not trow.empty:
            grade_val = trow.iloc[0].get("Grade", "")
            if grade_val is not None and str(grade_val).strip():
                return str(grade_val).strip()
    except Exception:
        pass
    return ""


def _get_scorecard_result_color(score_df: pd.DataFrame) -> str:
    """
    Returns the same color used by the scorecard top strip:
    Poor = red, Average/Soft = orange, Good/Stable = green, Very Good/Strong = darker green, Excellent = darkest green.
    """
    grade_val = _extract_total_grade_text(score_df)
    try:
        _verdict, color = _summary_grade_to_verdict_and_color(grade_val)
        return color
    except Exception:
        return SCORECARD_CENTER_BOX_BORDER


def _hex_to_rgba(hex_color: str, alpha: int = 255) -> tuple[int, int, int, int]:
    s = str(hex_color or "#000000").strip().lstrip("#")
    if len(s) == 3:
        s = "".join(ch * 2 for ch in s)
    try:
        r = int(s[0:2], 16)
        g = int(s[2:4], 16)
        b = int(s[4:6], 16)
    except Exception:
        r, g, b = 0, 0, 0
    return (r, g, b, int(alpha))


def _load_bold_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    font_candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf",
        "arialbd.ttf",
        "arial.ttf",
    ]
    for font_path in font_candidates:
        try:
            return ImageFont.truetype(font_path, int(size))
        except Exception:
            continue
    return ImageFont.load_default()


def _draw_center_score_on_clone(base_img: Image.Image, score_text: str, font_size: int, result_color: str | None = None) -> Image.Image:
    """
    Keeps the complete General scorecard image as-is, then draws one score
    in the center. Frame-to-frame zoom is created by increasing only this
    center score size, not by cropping the base image.
    """
    frame = base_img.convert("RGBA")
    w, h = frame.size

    score_text = str(score_text or "").strip()
    if not score_text:
        return frame.convert("RGB")

    label_text = str(SCORECARD_CENTER_SCORE_PREFIX or "")
    full_text = f"{label_text}{score_text}"

    font = _load_bold_font(font_size)
    draw = ImageDraw.Draw(frame)

    # Text bbox
    bbox = draw.textbbox((0, 0), full_text, font=font)
    text_w = bbox[2] - bbox[0]
    text_h = bbox[3] - bbox[1]

    box_w = max(int(w * SCORECARD_CENTER_BOX_WIDTH), text_w + int(font_size * 1.35))
    box_h = max(int(h * SCORECARD_CENTER_BOX_HEIGHT), text_h + int(font_size * 1.05))
    box_x1 = int((w - box_w) / 2)
    box_y1 = int((h * SCORECARD_CENTER_SCORE_Y) - (box_h / 2))
    box_x2 = box_x1 + box_w
    box_y2 = box_y1 + box_h

    overlay = Image.new("RGBA", frame.size, (0, 0, 0, 0))
    odraw = ImageDraw.Draw(overlay)
    active_result_color = str(result_color or SCORECARD_CENTER_BOX_BORDER).strip()
    active_label_color = active_result_color if SCORECARD_CENTER_USE_RESULT_COLOR else SCORECARD_CENTER_LABEL_COLOR
    active_score_color = active_result_color if SCORECARD_CENTER_USE_RESULT_COLOR else SCORECARD_CENTER_SCORE_COLOR
    active_border_color = active_result_color if SCORECARD_CENTER_USE_RESULT_COLOR else SCORECARD_CENTER_BOX_BORDER

    odraw.rounded_rectangle(
        (box_x1, box_y1, box_x2, box_y2),
        radius=int(SCORECARD_CENTER_BOX_RADIUS),
        fill=_hex_to_rgba(SCORECARD_CENTER_BOX_FILL, SCORECARD_CENTER_BOX_ALPHA),
        outline=_hex_to_rgba(active_border_color, 255),
        width=int(SCORECARD_CENTER_BOX_BORDER_WIDTH),
    )
    frame = Image.alpha_composite(frame, overlay)
    draw = ImageDraw.Draw(frame)

    # Draw label and score separately so the numeric score remains the focus.
    label_bbox = draw.textbbox((0, 0), label_text, font=font)
    score_bbox = draw.textbbox((0, 0), score_text, font=font)
    label_w = label_bbox[2] - label_bbox[0]
    score_w = score_bbox[2] - score_bbox[0]
    total_w = label_w + score_w
    x = int((w - total_w) / 2)
    y = int((box_y1 + box_y2 - text_h) / 2 - bbox[1])

    # Subtle shadow for readability.
    shadow_offset = max(2, int(font_size * 0.04))
    draw.text((x + shadow_offset, y + shadow_offset), label_text, font=font, fill=(0, 0, 0, 120))
    draw.text((x + label_w + shadow_offset, y + shadow_offset), score_text, font=font, fill=(0, 0, 0, 120))
    draw.text((x, y), label_text, font=font, fill=_hex_to_rgba(active_label_color, 255))
    draw.text((x + label_w, y), score_text, font=font, fill=_hex_to_rgba(active_score_color, 255))

    return frame.convert("RGB")


def generate_scorecard_video_frames(
    symbol: str,
    base_image_path: str,
    score_text: str = "",
    result_color: str | None = None,
    out_path_template: str | None = None,
    frame_group: str = "General",
) -> dict:
    """
    Generates 4 scorecard video frames from the given base image.

    Use this for both General and Instagram base scorecards so outputs are saved as:
        Image/VideoFrames/General/{SYMBOL}_Frame_01.jpeg
        Image/VideoFrames/Instagram/{SYMBOL}_Frame_01.jpeg
    """
    paths = {}

    if not GENERATE_SCORECARD_VIDEO_FRAMES:
        return paths

    group_name = str(frame_group or "General").strip()
    template = out_path_template or SCORECARD_VIDEO_FRAMES_OUTPUT_TEMPLATE

    if not base_image_path or not os.path.exists(base_image_path):
        print(f"⚠️ {group_name} scorecard video frames skipped. Base image not found: {base_image_path}")
        return paths

    try:
        base_img = Image.open(base_image_path).convert("RGB")
    except Exception as e:
        print(f"⚠️ {group_name} scorecard video frames skipped. Could not open base image: {e}")
        return paths

    font_sizes = list(SCORECARD_CENTER_SCORE_FONT_SIZES)
    if len(font_sizes) < SCORECARD_VIDEO_FRAMES_COUNT:
        last_size = int(font_sizes[-1]) if font_sizes else 46
        while len(font_sizes) < SCORECARD_VIDEO_FRAMES_COUNT:
            last_size += 12
            font_sizes.append(last_size)

    key_prefix = "Scorecard" + re.sub(r"[^A-Za-z0-9]+", "", group_name) + "VideoFrame"

    for idx in range(1, SCORECARD_VIDEO_FRAMES_COUNT + 1):
        font_size = int(font_sizes[idx - 1])
        frame = _draw_center_score_on_clone(
            base_img,
            score_text=score_text,
            font_size=font_size,
            result_color=result_color,
        )

        out_path = template.replace("{SYMBOL}", symbol).replace("{FRAME_NO}", f"{idx:02d}")
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        frame.save(out_path, quality=SCORECARD_VIDEO_FRAME_QUALITY, optimize=True)
        paths[f"{key_prefix}{idx:02d}"] = out_path
        print(f"Saved {group_name} Scorecard Video Frame {idx}: {out_path} | Center score font: {font_size}")

    return paths

def generate_all_summary_images(symbol: str, company_name: str, summary_payload: dict) -> dict:
    paths = {}
    if GENERATE_GENERAL_SUMMARY_IMAGE:
        paths["GeneralSummaryImage"] = export_result_summary_as_image(
            symbol=symbol,
            company_name=company_name,
            summary_payload=summary_payload,
            out_path_template=GENERAL_SUMMARY_IMAGE_OUTPUT_TEMPLATE,
            image_type="general",
        )
        print(f"Saved General Summary Image: {paths['GeneralSummaryImage']}")

    if GENERATE_INSTAGRAM_SUMMARY_IMAGE:
        paths["InstagramSummaryImage"] = export_result_summary_as_image(
            symbol=symbol,
            company_name=company_name,
            summary_payload=summary_payload,
            out_path_template=INSTAGRAM_SUMMARY_IMAGE_OUTPUT_TEMPLATE,
            image_type="instagram",
        )
        print(f"Saved Instagram Summary Image: {paths['InstagramSummaryImage']}")

    if GENERATE_STANDARD_SUMMARY_IMAGE:
        paths["StandardSummaryImage"] = export_result_summary_as_image(
            symbol=symbol,
            company_name=company_name,
            summary_payload=summary_payload,
            out_path_template=STANDARD_SUMMARY_IMAGE_OUTPUT_TEMPLATE,
            image_type="standard",
        )
        print(f"Saved Standard Summary Image: {paths['StandardSummaryImage']}")

    if GENERATE_REELS_SUMMARY_IMAGE:
        paths["ReelsSummaryImage"] = export_result_summary_as_image(
            symbol=symbol,
            company_name=company_name,
            summary_payload=summary_payload,
            out_path_template=REELS_SUMMARY_IMAGE_OUTPUT_TEMPLATE,
            image_type="reels",
        )
        print(f"Saved Reels Summary Image: {paths['ReelsSummaryImage']}")

    if GENERATE_REELS_FRAME_PACK:
        paths.update(generate_reels_frame_pack(symbol, company_name, summary_payload))

    return paths

def save_result_analysis_excel(symbol: str, company_name: str, score_df: pd.DataFrame, latest_quarter: str) -> str:
    quarter_safe = (latest_quarter or "Result").replace(" ", "_").replace("/", "_").replace("\\", "_").replace("-", "_")
    out_path = EXCEL_OUTPUT_TEMPLATE.replace("{SYMBOL}", symbol).replace("{QUARTER_SAFE}", quarter_safe)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        sheet = "ResultAnalysis"
        score_df.to_excel(writer, sheet_name=sheet, startrow=2, index=False)
        wb = writer.book
        ws = wb[sheet]

        title = f"{company_name} {latest_quarter} Results" if latest_quarter else f"{company_name} Results"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(score_df.columns)))
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=15, color="FFFFFF")
        ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="203864")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        header_fill = PatternFill("solid", fgColor="D9D9D9")
        total_fill = PatternFill("solid", fgColor="FFF2CC")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        header_row = 3
        for col in range(1, len(score_df.columns) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 18

        for r in range(header_row + 1, ws.max_row + 1):
            is_total = str(ws.cell(row=r, column=1).value).strip().upper() == "TOTAL"
            for c in range(1, len(score_df.columns) + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if is_total:
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
                if c in [3, 4, 5, 6, 7]:
                    cell.alignment = Alignment(horizontal="right")

    return out_path


def run_single_stock(symbol: str, company_name: str, screener_url: str) -> dict:
    clean_symbol = str(symbol).strip().upper().replace(".NS", "").replace(".BO", "")
    company_name = title_case_stock_name(company_name) or clean_symbol

    q = fetch_quarterly_from_screener(screener_url)
    fdf = build_features(q)
    score_df = score_quarter_parameters(fdf)

    latest_quarter = str(q["Quarter"].iloc[-1]).strip() if not q.empty else ""
    title_suffix = f"{latest_quarter} Results" if latest_quarter else "Results"

    impact_summary = {}
    try:
        impact_summary = calculate_result_impact_summary(q, clean_symbol)
    except Exception as e:
        print(f"⚠️ Result-impact summary calculation failed for {clean_symbol}: {e}")
        impact_summary = {
            "Direction": "N/A",
            "ExpectedMove1D%": "N/A",
            "ExpectedMove3D%": "N/A",
            "Confidence": "N/A",
        }

    summary_payload = build_result_summary_payload(fdf, score_df, latest_quarter, impact_summary)

    excel_path = ""
    if GENERATE_EXCEL_REPORT:
        excel_path = save_result_analysis_excel(clean_symbol, company_name, score_df, latest_quarter)
        print(f"Saved Excel: {excel_path}")

    general_img_path = ""
    instagram_img_path = ""

    if GENERATE_GENERAL_IMAGE:
        general_img_path = export_result_table_as_image(
            symbol=clean_symbol,
            company_name=company_name,
            score_df=score_df,
            total_row=None,
            out_path_template=GENERAL_IMAGE_OUTPUT_TEMPLATE,
            title_suffix=title_suffix,
            footer_text="",
            instagram_mode=False
        )
        print(f"Saved General Image: {general_img_path}")

    if GENERATE_INSTAGRAM_IMAGE:
        instagram_img_path = export_result_table_as_image(
            symbol=clean_symbol,
            company_name=company_name,
            score_df=score_df,
            total_row=None,
            out_path_template=INSTAGRAM_IMAGE_OUTPUT_TEMPLATE,
            title_suffix=title_suffix,
            footer_text="",
            instagram_mode=True
        )
        print(f"Saved Instagram Image: {instagram_img_path}")

    # Generate scorecard VideoFrames from both exact scorecard images.
    # General frames go to Image/VideoFrames/General/
    # Instagram frames go to Image/VideoFrames/Instagram/
    center_score_text = _extract_total_score_text(score_df)
    center_result_color = _get_scorecard_result_color(score_df)
    scorecard_video_frame_paths = {}

    if GENERATE_SCORECARD_GENERAL_VIDEO_FRAMES:
        scorecard_video_frame_paths.update(generate_scorecard_video_frames(
            clean_symbol,
            general_img_path,
            score_text=center_score_text,
            result_color=center_result_color,
            out_path_template=SCORECARD_GENERAL_VIDEO_FRAMES_OUTPUT_TEMPLATE,
            frame_group="General",
        ))

    if GENERATE_SCORECARD_INSTAGRAM_VIDEO_FRAMES:
        scorecard_video_frame_paths.update(generate_scorecard_video_frames(
            clean_symbol,
            instagram_img_path,
            score_text=center_score_text,
            result_color=center_result_color,
            out_path_template=SCORECARD_INSTAGRAM_VIDEO_FRAMES_OUTPUT_TEMPLATE,
            frame_group="Instagram",
        ))

    summary_image_paths = generate_all_summary_images(clean_symbol, company_name, summary_payload)

    return {
        "Symbol": clean_symbol,
        "CompanyName": company_name,
        "ScreenerURL": screener_url,
        "LatestQuarter": latest_quarter,
        "Status": "SUCCESS",
        "Excel": excel_path,
        "GeneralImage": general_img_path,
        "InstagramImage": instagram_img_path,
        **scorecard_video_frame_paths,
        **summary_image_paths,
    }

def _parse_cli_args():
    """
    DATE filled  -> batch mode from ResultCalendar Excel.
    DATE blank   -> single-symbol mode using:
                    python .\\ResultImpactPredictor_v6.py --symbol BAJAJHLDNG

    This keeps the existing DATE-based functionality unchanged while allowing
    symbol-only image generation when DATE is blank.
    """
    parser = argparse.ArgumentParser(description="Result impact image generator")
    parser.add_argument(
        "--symbol",
        type=str,
        default="",
        help="Single stock symbol, e.g. BAJAJHLDNG. Used mainly when DATE is blank."
    )
    parser.add_argument(
        "--company-name",
        type=str,
        default="",
        help="Optional display name for the title. If blank, symbol is used."
    )
    parser.add_argument(
        "--screener-url",
        type=str,
        default="",
        help="Optional Screener URL. If blank, it is generated from --symbol."
    )
    return parser.parse_args()


def main():
    args = _parse_cli_args()

    date_text = str(DATE or "").strip()
    cli_symbol = str(args.symbol or "").strip().upper().replace(".NS", "").replace(".BO", "")

    # =====================================================
    # CASE 1: DATE is blank, but --symbol is provided
    # Example:
    # python .\ResultImpactPredictor_v6.py --symbol BAJAJHLDNG
    # =====================================================
    if not date_text:
        if not cli_symbol:
            raise ValueError(
                "DATE is blank, so run with --symbol. "
                "Example: python .\\ResultImpactPredictor_v6.py --symbol BAJAJHLDNG"
            )

        screener_url = str(args.screener_url).strip() or build_screener_url_from_symbol(cli_symbol)

        if str(args.company_name).strip():
            company_name = title_case_stock_name(args.company_name)
        else:
            company_name = fetch_company_name_from_screener(screener_url, cli_symbol) or cli_symbol

        print("\n======================================================")
        print("Single-symbol mode enabled because DATE is blank.")
        print(f"Symbol   : {cli_symbol}")
        print(f"Company  : {company_name}")
        print(f"Screener : {screener_url}")
        print("======================================================\n")

        result = run_single_stock(cli_symbol, company_name, screener_url)

        os.makedirs("Result", exist_ok=True)
        log_file = f"Result/Single_Result_Generation_Log_{cli_symbol}.csv"
        pd.DataFrame([result]).to_csv(log_file, index=False)

        print("\n======================================================")
        print("SINGLE RESULT IMAGE GENERATION COMPLETED")
        print(f"Symbol saved : {cli_symbol}")
        print(f"Log saved    : {log_file}")
        print("======================================================\n")
        return

    # =====================================================
    # CASE 2: Existing DATE-based batch mode
    # Existing functionality remains unchanged.
    # =====================================================
    print("\n======================================================")
    print(f"DATE automation enabled: {date_text}")
    print("Reading stocks from local ResultCalendar Excel only.")
    print("No Moneycontrol, NSE, Selenium, Yahoo, or dynamic symbol mapping logic is used.")
    print("======================================================\n")

    stock_df = read_result_calendar_stocklist_excel(date_text, StockList)
    logs = []

    for i, row in stock_df.iterrows():
        symbol = str(row["Symbol"]).strip().upper()
        company_name = str(row["CompanyName"]).strip()
        screener_url = str(row["ScreenerURL"]).strip()

        print("\n------------------------------------------------------")
        print(f"Processing {i + 1}/{len(stock_df)}: {company_name} ({symbol})")
        print(f"Screener: {screener_url}")
        print("------------------------------------------------------")

        try:
            logs.append(run_single_stock(symbol, company_name, screener_url))
        except Exception as e:
            err = str(e)
            print(f"❌ Failed for {symbol}: {err}")
            logs.append({
                "Symbol": symbol,
                "CompanyName": company_name,
                "ScreenerURL": screener_url,
                "Status": "FAILED",
                "Error": err,
            })
            if not CONTINUE_ON_STOCK_ERROR:
                raise

    os.makedirs("Result", exist_ok=True)
    log_file = f"Result/Auto_Result_Generation_Log_{parse_user_date_to_dd_mm_yyyy(date_text)}.csv"
    pd.DataFrame(logs).to_csv(log_file, index=False)

    print("\n======================================================")
    print("AUTO RESULT IMAGE GENERATION COMPLETED")
    print(f"Total valid stocks : {len(stock_df)}")
    print(f"Log saved          : {log_file}")
    print("======================================================\n")


if __name__ == "__main__":
    main()
